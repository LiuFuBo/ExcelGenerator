#!/usr/bin/env python3
"""
Excel订单生成器 - 跨平台GUI应用 (PySide6版本)
功能：
1. 输入Excel名称
2. 输入销售日期总条数
3. 输入销售单号总条数
4. 自动生成销售日期（当月1号到今日，9:00-21:00）
5. 自动生成销售单号（S+YYMMDD+14位1开头随机数）
6. 错误日志显示
7. 输入校验
"""

import sys
import os
import json
import random
import calendar
import traceback
from datetime import datetime, timedelta

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTextEdit, QMessageBox, QFileDialog,
    QStackedWidget, QFrame, QScrollArea, QSizePolicy,
    QCheckBox, QGridLayout, QDialog, QGraphicsOpacityEffect,
    QRadioButton, QButtonGroup
)
from PySide6.QtCore import Qt, QSize, QRect, QTimer, QStandardPaths, QPropertyAnimation
from PySide6.QtGui import QFont, QColor, QIcon, QPixmap, QImage, QPainter, QPen, QConicalGradient

from PIL import Image, ImageOps, ImageFont, ImageDraw, ImageFilter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ─── macOS-style QSS ───────────────────────────────────────────────────────────

MACOS_QSS = """
/* Global */
QWidget {
    font-family: -apple-system, 'SF Pro Text', 'Helvetica Neue', sans-serif;
    color: #1d1d1f;
}

/* ── Sidebar ── */
#sidebar {
    background: rgba(232,232,237,.92);
    border-right: 1px solid #e5e5ea;
    min-width: 200px;
    max-width: 200px;
}
#sidebar_label {
    font-size: 11px;
    color: #aeaeb2;
    font-weight: 600;
    padding: 18px 0px 6px 0px;
}
#sidebar_item {
    font-size: 13px;
    color: #6e6e73;
    font-weight: 500;
    padding: 7px 0px;
    border-radius: 6px;
    margin-bottom: 2px;
    text-align: left;
}
#sidebar_item:hover {
    background: rgba(0,0,0,.04);
}
#sidebar_item_active {
    font-size: 13px;
    color: #0071e3;
    font-weight: 500;
    padding: 7px 0px;
    border-radius: 6px;
    margin-bottom: 2px;
    background: rgba(0,102,255,.08);
    text-align: left;
}
#sidebar_item_icon {
    font-size: 15px;
    min-width: 20px;
}

/* ── Form Card ── */
#form_card {
    background: #ffffff;
    border-radius: 12px;
    border: 1px solid rgba(0,0,0,.04);
    padding: 18px 22px 23px 22px;
}

/* ── Field Title ── */
#field_title {
    font-size: 13px;
    font-weight: 600;
    color: #1d1d1f;
    margin-bottom: 5px;
}
#field_title_icon {
    font-size: 13px;
    opacity: .7;
}

/* ── All Line Inputs ── */
#excel_name_input, #month_input, #date_count_input, #order_count_input, #pos_count_input, #filepath_display {
    background: #fafafa;
    border: 1px solid #cdcdcd;
    border-radius: 5px;
    padding: 5px 12px;
    font-size: 14px;
    color: #1d1d1f;
    min-height: 25px;
    max-height: 25px;
    selection-background-color: rgba(0,113,227,.2);
}
#excel_name_input:focus, #month_input:focus, #date_count_input:focus, #order_count_input:focus, #pos_count_input:focus {
    border-color: #0071e3;
    background: #ffffff;
}
#excel_name_input:hover:!focus, #month_input:hover:!focus, #date_count_input:hover:!focus, #order_count_input:hover:!focus, #pos_count_input:hover:!focus, #filepath_display:hover {
    border-color: #b0b0b6;
}
#excel_name_input:disabled {
    background: #e8e8e8;
    color: #999;
}
#filepath_display {
    border: 1px solid #cdcdcd;
    font-size: 13px;
    color: #aeaeb2;
}

/* ── Primary Button (选择文件) ── */
#file_select_btn {
    background: #e0e0e0;
    color: #666;
    border: none;
    border-radius: 5px;
    font-size: 13px;
    font-weight: 500;
}
#file_select_btn:hover {
    background: #d0d0d0;
}

/* ── Secondary Button (清除) ── */
.mac-btn-secondary {
    background: transparent;
    color: #0071e3;
    border: 1px solid rgba(0,113,227,.2);
    border-radius: 8px;
    padding: 5px 12px;
    font-size: 12px;
    font-weight: 400;
    min-height: 34px;
}
.mac-btn-secondary:hover {
    background: rgba(0,113,227,.06);
    border-color: rgba(0,113,227,.3);
}

/* ── Generate Button ── */
#generate_btn {
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #0077ed, stop:1 #0066d6);
    color: #ffffff;
    border: none;
    border-radius: 5px;
    padding: 5px 24px;
    font-size: 15px;
    font-weight: 500;
    min-height: 35px;
    max-height: 35px;
}
#generate_btn:hover {
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #006ee0, stop:1 #005bc8);
}

/* ── Log Header ── */
#log_title {
    font-size: 13px;
    font-weight: 600;
    color: #1d1d1f;
}
#log_title_icon {
    font-size: 13px;
    opacity: .7;
}

/* ── Copy Log Button ── */
#copy_log_btn {
    background: transparent;
    color: #0071e3;
    border: 1px solid rgba(0,113,227,.2);
    border-radius: 6px;
    padding: 3px 10px;
    font-size: 12px;
    font-weight: 400;
    min-height: 19px;
}
#copy_log_btn:hover {
    background: rgba(0,113,227,.06);
    border-color: rgba(0,113,227,.3);
}

/* ── Log Box ── */
#log_area {
    background: #1d1d24;
    border: 1px solid #2c2c34;
    border-radius: 8px;
    padding: 12px 14px;
    font-family: 'SF Mono', 'Menlo', 'Monaco', monospace;
    font-size: 12px;
    color: #8e8e93;
    selection-background-color: rgba(0,113,227,.3);
}

/* ── Month Hint ── */
#month_hint {
    font-size: 13px;
    color: #0071e3;
    font-weight: 400;
}

/* ── Placeholder page ── */
#placeholder_icon {
    font-size: 48px;
    opacity: .3;
}
#placeholder_text {
    font-size: 14px;
    color: #aeaeb2;
}

/* ── Watermark Page ── */
#wm_image_card {
    background: #ffffff;
    border-radius: 10px;
    border: 1px solid rgba(0,0,0,.04);
    padding: 14px 16px;
}
#wm_image_title {
    font-size: 14px;
    font-weight: 600;
    color: #1d1d1f;
}
#wm_select_btn {
    background: #3B82F6;
    color: #ffffff;
    border: none;
    border-radius: 6px;
    padding: 6px 14px;
    font-size: 13px;
    font-weight: 500;
}
#wm_select_btn:hover { background: #2563EB; }
#wm_clear_btn {
    background: transparent;
    color: #6e6e73;
    border: 1px solid #cdcdcd;
    border-radius: 6px;
    padding: 4px 10px;
    font-size: 12px;
}
#wm_clear_btn:hover { background: rgba(0,0,0,.04); }
#wm_image_scroll {
    background: transparent;
    border: none;
}
#wm_image_scroll QScrollBar:vertical {
    width: 6px;
    background: transparent;
    border-radius: 3px;
}
#wm_image_scroll QScrollBar::handle:vertical {
    background: #CBD5E1;
    border-radius: 3px;
    min-height: 30px;
}
#wm_image_scroll QScrollBar::add-line:vertical, #wm_image_scroll QScrollBar::sub-line:vertical {
    height: 0px;
}
#wm_image_item {
    background: #F8FAFC;
    border: 2px solid #E2E8F0;
    border-radius: 8px;
}
#wm_image_item:hover { border-color: #CBD5E1; }
#wm_img_btn {
    background: transparent;
    border: none;
    padding: 0px;
    margin: 0px;
}
#wm_img_btn:hover { background: transparent; }
#wm_filename_label {
    font-size: 10px;
    color: #6e6e73;
}
#wm_delete_btn {
    background: rgba(239,68,68,.15);
    color: #EF4444;
    border: none;
    border-radius: 4px;
    font-size: 9px;
    min-width: 18px;
    max-width: 18px;
    min-height: 18px;
    max-height: 18px;
    padding: 0px;
}
#wm_delete_btn:hover { background: rgba(239,68,68,.3); }
#wm_skeleton {
    background: #F8FAFC;
    border: 2px solid #E2E8F0;
    border-radius: 8px;
}
#wm_skeleton_shimmer {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #E2E8F0, stop:0.5 #F1F5F9, stop:1 #E2E8F0);
    border-radius: 4px;
    min-height: 60px;
}
#wm_empty_icon {
    font-size: 36px;
    opacity: .35;
}
#wm_empty_text {
    font-size: 13px;
    color: #94A3B8;
}
#wm_empty_hint {
    font-size: 11px;
    color: #CBD5E1;
}
#wm_city_card {
    background: #ffffff;
    border-radius: 10px;
    border: 1px solid rgba(0,0,0,.04);
    padding: 16px;
}
#wm_city_title {
    font-size: 14px;
    font-weight: 600;
    color: #1d1d1f;
}
#wm_city_grid {
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    padding: 12px;
}
#wm_city_checkbox {
    font-size: 13px;
    color: #1d1d1f;
    spacing: 6px;
}
#wm_city_checkbox::indicator {
    width: 18px;
    height: 18px;
    border-radius: 4px;
    border: 1.5px solid #cdcdcd;
    background: #ffffff;
}
#wm_city_checkbox::indicator:checked {
    background: #3B82F6;
    border-color: #3B82F6;
    image: url(__CHECKMARK_ICON__);
}
#wm_all_city_checkbox {
    font-size: 13px;
    font-weight: 600;
    color: #3B82F6;
    spacing: 6px;
}
#wm_all_city_checkbox::indicator {
    width: 18px;
    height: 18px;
    border-radius: 4px;
    border: 1.5px solid #3B82F6;
    background: #ffffff;
}
#wm_all_city_checkbox::indicator:checked {
    background: #3B82F6;
    border-color: #3B82F6;
    image: url(__CHECKMARK_ICON__);
}
#wm_date_card {
    background: #ffffff;
    border-radius: 10px;
    border: 1px solid rgba(0,0,0,.04);
    padding: 16px;
}
#wm_date_title {
    font-size: 14px;
    font-weight: 600;
    color: #1d1d1f;
}
#wm_date_label {
    font-size: 13px;
    color: #6e6e73;
}
#wm_year_input, #wm_month_input {
    background: #F8FAFC;
    border: 1px solid #cdcdcd;
    border-radius: 6px;
    padding: 5px 10px;
    font-size: 14px;
    color: #1d1d1f;
    min-height: 32px;
    max-height: 32px;
}
#wm_year_input:focus, #wm_month_input:focus {
    border-color: #3B82F6;
}
#wm_start_btn {
    background: #3B82F6;
    color: #ffffff;
    border: none;
    border-radius: 8px;
    padding: 0px;
    font-size: 14px;
    font-weight: 600;
    min-height: 40px;
    max-height: 40px;
    min-width: 220px;
    max-width: 220px;
}
#wm_start_btn:hover { background: #2563EB; }

/* ── Image Picker Dialog ── */
#wm_picker_path_label {
    font-size: 13px;
    color: #1d1d1f;
    padding: 5px 10px;
}
#wm_picker_folder_btn {
    background: #3B82F6;
    color: #ffffff;
    border: none;
    border-radius: 6px;
    padding: 6px 14px;
    font-size: 13px;
    font-weight: 500;
}
#wm_picker_folder_btn:hover { background: #2563EB; }
#wm_picker_selectall_btn, #wm_picker_deselect_btn {
    background: transparent;
    color: #6e6e73;
    border: 1px solid #cdcdcd;
    border-radius: 6px;
    padding: 4px 10px;
    font-size: 12px;
}
#wm_picker_selectall_btn:hover, #wm_picker_deselect_btn:hover { background: rgba(0,0,0,.04); }
#wm_picker_count_label {
    font-size: 12px;
    color: #94A3B8;
}
#wm_picker_item {
    background: #F8FAFC;
    border: 2px solid #E2E8F0;
    border-radius: 8px;
}
#wm_picker_item_cb {
    font-size: 12px;
    color: #1d1d1f;
    spacing: 4px;
}
#wm_picker_item_cb::indicator {
    width: 18px;
    height: 18px;
    border-radius: 4px;
    border: 1.5px solid #cdcdcd;
    background: #ffffff;
}
#wm_picker_item_cb::indicator:checked {
    background: #3B82F6;
    border-color: #3B82F6;
    image: url(__CHECKMARK_ICON__);
}
#wm_picker_img_btn {
    background: transparent;
    border: none;
    padding: 0px;
    margin: 0px;
}
#wm_picker_filename {
    font-size: 10px;
    color: #6e6e73;
}
#wm_picker_skeleton {
    background: #F8FAFC;
    border: 2px solid #E2E8F0;
    border-radius: 8px;
}
#wm_skeleton_shimmer {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #E2E8F0, stop:0.5 #F1F5F9, stop:1 #E2E8F0);
    border-radius: 4px;
    min-height: 60px;
}
#wm_skeleton_text {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #E2E8F0, stop:0.5 #F1F5F9, stop:1 #E2E8F0);
    border-radius: 3px;
    min-height: 12px;
    max-height: 12px;
}
#wm_skeleton_checkbox {
    background: #E2E8F0;
    border-radius: 4px;
}
#wm_picker_confirm_btn {
    background: #3B82F6;
    color: #ffffff;
    border: none;
    border-radius: 8px;
    padding: 0px;
    font-size: 14px;
    font-weight: 600;
    min-height: 40px;
    max-height: 40px;
    min-width: 220px;
    max-width: 220px;
}
#wm_picker_confirm_btn:hover { background: #2563EB; }
#wm_picker_cancel_btn {
    background: transparent;
    color: #6e6e73;
    border: 1px solid #cdcdcd;
    border-radius: 8px;
    padding: 0px;
    font-size: 14px;
    font-weight: 500;
    min-height: 40px;
    max-height: 40px;
    min-width: 120px;
    max-width: 120px;
}
#wm_picker_cancel_btn:hover { background: rgba(0,0,0,.04); }

/* ── Compress Page ── */
#cp_image_card {
    background: #ffffff;
    border-radius: 10px;
    border: 1px solid rgba(0,0,0,.04);
    padding: 14px 16px;
}
#cp_image_title {
    font-size: 14px;
    font-weight: 600;
    color: #1d1d1f;
}
#cp_select_btn {
    background: #3B82F6;
    color: #ffffff;
    border: none;
    border-radius: 6px;
    padding: 6px 14px;
    font-size: 13px;
    font-weight: 500;
}
#cp_select_btn:hover { background: #2563EB; }
#cp_clear_btn {
    background: transparent;
    color: #6e6e73;
    border: 1px solid #E2E8F0;
    border-radius: 4px;
    padding: 4px 10px;
    font-size: 12px;
}
#cp_clear_btn:hover { background: rgba(0,0,0,.03); }
#cp_image_scroll {
    background: transparent;
    border: none;
}
#cp_image_scroll QScrollBar:vertical {
    width: 6px;
    background: transparent;
    border-radius: 3px;
}
#cp_image_scroll QScrollBar::handle:vertical {
    background: #CBD5E1;
    border-radius: 3px;
    min-height: 30px;
}
#cp_image_scroll QScrollBar::add-line:vertical, #cp_image_scroll QScrollBar::sub-line:vertical {
    height: 0px;
}
#cp_image_item {
    background: #F8FAFC;
    border: 2px solid #E2E8F0;
    border-radius: 8px;
}
#cp_image_item:hover { border-color: #CBD5E1; }
#cp_img_btn {
    background: transparent;
    border: none;
    padding: 0px;
    margin: 0px;
}
#cp_img_btn:hover { background: transparent; }
#cp_filename_label {
    font-size: 10px;
    color: #6e6e73;
}
#cp_delete_btn {
    background: rgba(239,68,68,.15);
    color: #EF4444;
    border: none;
    border-radius: 4px;
    font-size: 9px;
}
#cp_delete_btn:hover { background: rgba(239,68,68,.25); }
#cp_skeleton {
    background: #F8FAFC;
    border: 2px solid #E2E8F0;
    border-radius: 8px;
}
#cp_skeleton_shimmer {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #E2E8F0, stop:0.5 #F1F5F9, stop:1 #E2E8F0);
    border-radius: 4px;
    min-height: 60px;
}
#cp_empty_icon {
    font-size: 36px;
    opacity: .35;
}
#cp_empty_text {
    font-size: 13px;
    color: #94A3B8;
}
#cp_empty_hint {
    font-size: 11px;
    color: #CBD5E1;
}
#cp_stats_card {
    background: #ffffff;
    border-radius: 10px;
    border: 1px solid rgba(0,0,0,.04);
    padding: 16px;
}
#cp_stats_title {
    font-size: 14px;
    font-weight: 600;
    color: #1E293B;
}
#cp_stat_item {
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    padding: 12px 14px;
}
#cp_stat_label {
    font-size: 12px;
    color: #64748B;
}
#cp_stat_value {
    font-size: 20px;
    font-weight: 700;
    color: #1E293B;
}
#cp_stat_value_orange {
    font-size: 20px;
    font-weight: 700;
    color: #F59E0B;
}
#cp_stat_value_green {
    font-size: 20px;
    font-weight: 700;
    color: #22C55E;
}
#cp_stat_sub {
    font-size: 11px;
    color: #94A3B8;
}
#cp_start_btn {
    background: #3B82F6;
    color: #ffffff;
    border: none;
    border-radius: 8px;
    padding: 10px 0px;
    font-size: 14px;
    font-weight: 600;
    width: 220px;
}
#cp_start_btn:hover { background: #2563EB; }
#cp_method_card {
    background: #ffffff;
    border-radius: 10px;
}
#cp_method_title {
    font-size: 14px;
    font-weight: 600;
    color: #1d1d1f;
    padding-bottom: 4px;
}
#cp_method_radio {
    font-size: 13px;
    color: #1d1d1f;
    spacing: 6px;
    padding: 8px 12px;
}
#cp_method_radio::indicator {
    width: 10px;
    height: 10px;
    border-radius: 5px;
    border: 1px solid #CBD5E1;
    background: #ffffff;
}
#cp_method_radio::indicator:checked {
    border-color: #3B82F6;
    background: #3B82F6;
}
#cp_method_note {
    font-size: 11px;
    color: #94A3B8;
    padding-left: 28px;
}
"""


class ImagePickerDialog(QDialog):
    def __init__(self, parent, existing_paths):
        super().__init__(parent)
        self.setWindowTitle("选择图片")
        self.setMinimumSize(750, 500)
        self.resize(750, 500)
        self.existing_paths = set(existing_paths)
        self.folder_path = ""
        self.image_files = []
        self.checkboxes = []
        self._parent = parent
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # ── Folder row ──
        folder_row = QWidget()
        folder_layout = QHBoxLayout(folder_row)
        folder_layout.setContentsMargins(0, 0, 0, 0)

        folder_icon = QLabel("📁")
        folder_icon.setStyleSheet("font-size: 16px;")
        folder_layout.addWidget(folder_icon)

        self.path_label = QLabel("请选择图片文件夹")
        self.path_label.setObjectName("wm_picker_path_label")
        folder_layout.addWidget(self.path_label, 1)

        folder_btn = QPushButton("选择文件夹")
        folder_btn.setObjectName("wm_picker_folder_btn")
        folder_btn.setCursor(Qt.PointingHandCursor)
        folder_btn.clicked.connect(self._on_select_folder)
        folder_layout.addWidget(folder_btn)

        layout.addWidget(folder_row)

        # ── Toolbar row ──
        toolbar = QWidget()
        tb_layout = QHBoxLayout(toolbar)
        tb_layout.setContentsMargins(0, 0, 0, 0)

        selectall_btn = QPushButton("全选")
        selectall_btn.setObjectName("wm_picker_selectall_btn")
        selectall_btn.setCursor(Qt.PointingHandCursor)
        selectall_btn.clicked.connect(self._on_select_all)
        tb_layout.addWidget(selectall_btn)

        deselect_btn = QPushButton("取消全选")
        deselect_btn.setObjectName("wm_picker_deselect_btn")
        deselect_btn.setCursor(Qt.PointingHandCursor)
        deselect_btn.clicked.connect(self._on_deselect_all)
        tb_layout.addWidget(deselect_btn)

        tb_layout.addStretch()

        self.count_label = QLabel("已选 0/0 张")
        self.count_label.setObjectName("wm_picker_count_label")
        tb_layout.addWidget(self.count_label)

        layout.addWidget(toolbar)

        # ── Scrollable grid ──
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")

        self.grid_container = QWidget()
        self.grid_container.setStyleSheet("background: transparent;")
        self.grid_layout = QGridLayout(self.grid_container)
        self.grid_layout.setSpacing(10)
        self.grid_layout.setContentsMargins(0, 0, 0, 0)

        # Empty hint
        self.empty_hint = QLabel("请先选择文件夹，图片将显示在此处")
        self.empty_hint.setAlignment(Qt.AlignCenter)
        self.empty_hint.setStyleSheet("font-size: 13px; color: #94A3B8; padding: 40px;")
        self.grid_layout.addWidget(self.empty_hint, 0, 0)

        scroll.setWidget(self.grid_container)
        layout.addWidget(scroll, 1)

        # ── Bottom buttons ──
        btn_row = QWidget()
        btn_layout = QHBoxLayout(btn_row)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        btn_layout.addStretch()

        confirm_btn = QPushButton("确认添加")
        confirm_btn.setObjectName("wm_picker_confirm_btn")
        confirm_btn.setCursor(Qt.PointingHandCursor)
        confirm_btn.clicked.connect(self._on_confirm)
        btn_layout.addWidget(confirm_btn)

        btn_layout.addSpacing(12)

        cancel_btn = QPushButton("取消")
        cancel_btn.setObjectName("wm_picker_cancel_btn")
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        btn_layout.addStretch()
        layout.addWidget(btn_row)

    def _on_select_folder(self):
        dir_path = QFileDialog.getExistingDirectory(self, "选择图片文件夹")
        if dir_path:
            self.folder_path = dir_path
            self.path_label.setText(dir_path)
            self._scan_folder()

    def _scan_folder(self):
        exts = ('.jpg', '.jpeg', '.png', '.webp')
        self.image_files = []
        for fname in sorted(os.listdir(self.folder_path)):
            if fname.lower().endswith(exts):
                full_path = os.path.join(self.folder_path, fname)
                self.image_files.append(full_path)
        self._clear_grid()
        self._load_index = 0

        if not self.image_files:
            self.empty_hint = QLabel("文件夹中没有图片文件")
            self.empty_hint.setAlignment(Qt.AlignCenter)
            self.empty_hint.setStyleSheet("font-size: 13px; color: #94A3B8; padding: 40px;")
            self.grid_layout.addWidget(self.empty_hint, 0, 0)
            self._update_count()
            return

        dialog_grid_width = self.grid_container.width()
        if dialog_grid_width < 100:
            dialog_grid_width = self.width() - 32
        picker_spacing = 10 * 3
        picker_padding = 8 * 4
        self._picker_thumb_width = max(80, (dialog_grid_width - picker_spacing - picker_padding) // 4)
        tw = self._picker_thumb_width

        # Create skeleton cards for all images immediately
        self._skeleton_cards = {}
        for i in range(len(self.image_files)):
            row, col = i // 4, i % 4
            skeleton = self._create_skeleton_card(tw)
            self._skeleton_cards[i] = skeleton
            self.grid_layout.addWidget(skeleton, row, col)

        # Start replacing skeletons with real images after brief render delay
        QTimer.singleShot(50, self._load_next_image)

    def _clear_grid(self):
        while self.grid_layout.count():
            item = self.grid_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.checkboxes.clear()

    def _create_skeleton_card(self, tw):
        card = QFrame()
        card.setObjectName("wm_picker_skeleton")
        card.setMinimumHeight(180)
        card.setMaximumWidth(tw + 8)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(4, 15, 4, 15)
        card_layout.setSpacing(4)

        # Fake checkbox row
        top_row = QWidget()
        top_layout = QHBoxLayout(top_row)
        top_layout.setContentsMargins(0, 0, 0, 0)
        fake_cb = QLabel()
        fake_cb.setObjectName("wm_skeleton_checkbox")
        fake_cb.setFixedSize(18, 18)
        top_layout.addWidget(fake_cb)
        fake_text = QLabel()
        fake_text.setObjectName("wm_skeleton_text")
        fake_text.setMinimumWidth(60)
        top_layout.addWidget(fake_text)
        top_layout.addStretch()
        card_layout.addWidget(top_row)

        # Fake thumbnail area
        shimmer = QLabel()
        shimmer.setObjectName("wm_skeleton_shimmer")
        shimmer.setFixedSize(tw, 200)
        card_layout.addWidget(shimmer)

        return card

    def _load_next_image(self):
        if self._load_index >= len(self.image_files):
            return

        path = self.image_files[self._load_index]
        index = self._load_index
        row = index // 4
        col = index % 4
        tw = self._picker_thumb_width

        # Remove skeleton card at this position
        skeleton = self._skeleton_cards.get(index)
        if skeleton:
            self.grid_layout.removeWidget(skeleton)
            skeleton.deleteLater()

        card = QFrame()
        card.setObjectName("wm_picker_item")
        card.setMinimumHeight(180)
        card.setMaximumWidth(tw + 8)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(4, 15, 4, 15)
        card_layout.setSpacing(4)

        top_row = QWidget()
        top_layout = QHBoxLayout(top_row)
        top_layout.setContentsMargins(0, 0, 0, 0)

        cb = QCheckBox(os.path.basename(path))
        cb.setObjectName("wm_picker_item_cb")
        cb.setCursor(Qt.PointingHandCursor)
        is_existing = path in self.existing_paths
        if is_existing:
            cb.setEnabled(False)
            cb.setChecked(True)
            cb.setText(os.path.basename(path) + " (已添加)")
        cb.stateChanged.connect(self._update_count)
        top_layout.addWidget(cb)
        card_layout.addWidget(top_row)

        self.checkboxes.append(cb)

        thumb_btn = QPushButton()
        thumb_btn.setObjectName("wm_picker_img_btn")
        thumb_btn.setCursor(Qt.PointingHandCursor)
        pixmap = self._parent._load_rotated_pixmap(path, target_width=tw)
        if not pixmap.isNull():
            scaled = pixmap.scaledToWidth(tw, Qt.SmoothTransformation)
            if scaled.height() > 200:
                scaled = pixmap.scaled(tw, 200, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
                # Crop center: take tw×200 from the scaled image
                crop_y = (scaled.height() - 200) // 2
                scaled = scaled.copy(0, crop_y, tw, 200)
            thumb_btn.setIcon(QIcon(scaled))
            thumb_btn.setIconSize(QSize(tw, scaled.height()))
            thumb_btn.setFixedSize(tw, scaled.height())
        else:
            thumb_btn.setText("无法加载")
            thumb_btn.setFixedSize(tw, 200)
        thumb_btn.clicked.connect(lambda checked=False, i=index: self._on_toggle_select(i))
        card_layout.addWidget(thumb_btn, 1)

        self.grid_layout.addWidget(card, row, col)
        self._load_index += 1
        self._update_count()
        QTimer.singleShot(0, self._load_next_image)

    def _on_toggle_select(self, index):
        cb = self.checkboxes[index]
        if cb.isEnabled():
            cb.setChecked(not cb.isChecked())

    def _on_select_all(self):
        for cb in self.checkboxes:
            if cb.isEnabled():
                cb.setChecked(True)
        self._update_count()

    def _on_deselect_all(self):
        for cb in self.checkboxes:
            if cb.isEnabled():
                cb.setChecked(False)
        self._update_count()

    def _update_count(self):
        total = len(self.checkboxes)
        selected = sum(1 for cb in self.checkboxes if cb.isChecked())
        self.count_label.setText(f"已选 {selected}/{total} 张")

    def _on_confirm(self):
        selected_paths = []
        for i, cb in enumerate(self.checkboxes):
            if cb.isChecked():
                selected_paths.append(self.image_files[i])
        self.selected_paths = selected_paths
        self.accept()

    def get_selected_paths(self):
        return getattr(self, 'selected_paths', [])


class LoadingOverlay(QWidget):
    """半透明遮罩 + 转圈动画 + 成功提示"""
    def __init__(self, parent, loading_text="正在添加水印..."):
        super().__init__(parent)
        self.angle = 0
        self.success_visible = False
        self.loading_text = loading_text
        self.setObjectName("loading_overlay")
        self.setStyleSheet("""
            #loading_overlay { background: rgba(0,0,0,0.45); }
        """)
        # Spinner
        self.spinner_size = 48
        self.spinner_pixmap = None
        # Timer for rotation
        self.timer = QTimer(self)
        self.timer.timeout.connect(self._tick)
        # Success auto-hide timer
        self.success_timer = QTimer(self)
        self.success_timer.setSingleShot(True)
        self.success_timer.timeout.connect(self.hide)

    def show_loading(self):
        self.success_visible = False
        if self.parent():
            self.setGeometry(self.parent().rect())
        self.update()
        self.show()
        self.raise_()
        self.timer.start(40)  # ~25fps

    def show_success(self, msg="水印相册已生成！"):
        self.timer.stop()
        self.success_visible = True
        self.success_msg = msg
        self.update()
        self.success_timer.start(2000)

    def _tick(self):
        self.angle = (self.angle + 12) % 360
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        # Semi-transparent background
        painter.fillRect(self.rect(), QColor(0, 0, 0, 115))

        if self.success_visible:
            # Success message
            painter.setPen(QColor(255, 255, 255))
            font = QFont("PingFang SC", 20)
            font.setWeight(QFont.Bold)
            painter.setFont(font)
            painter.drawText(self.rect(), Qt.AlignCenter, self.success_msg)
        else:
            # Spinner + text (centered, spinner offset 80pt up from center)
            cx = self.width() // 2
            cy = self.height() // 2 - 80
            r = 24
            # Draw arc segments for spinner
            pen = QPen(QColor(226, 232, 240, 200), 3)
            pen.setCapStyle(Qt.RoundCap)
            painter.setPen(pen)
            painter.drawArc(cx - r, cy - r, 2*r, 2*r, self.angle * 16, 270 * 16)
            # Brighter leading edge
            pen2 = QPen(QColor(255, 255, 255), 3)
            pen2.setCapStyle(Qt.RoundCap)
            painter.setPen(pen2)
            painter.drawArc(cx - r, cy - r, 2*r, 2*r, self.angle * 16, 30 * 16)
            # Text below spinner
            painter.setPen(QColor(226, 232, 240))
            font = QFont("PingFang SC", 16)
            painter.setFont(font)
            text_rect = QRect(cx - 150, cy + r + 20, 300, 40)
            painter.drawText(text_rect, Qt.AlignCenter, self.loading_text)

    def resizeEvent(self, event):
        super().resizeEvent(event)


class ExcelGeneratorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("许愿瓶")
        screen = QApplication.primaryScreen()
        screen_height = screen.availableGeometry().height()
        min_height = int(screen_height * 2 / 3)
        default_height = screen_height - 20
        self.setMinimumSize(1000, min_height)
        self.resize(1000, default_height)
        self.selected_file_path = ""
        self._pending_logs = []
        self._post_log_callback = None
        self._wm_pending_logs = []
        self._wm_post_log_callback = None
        self.wm_image_paths = []
        self.wm_image_items = {}
        self._wm_skeleton_cards = {}
        self._wm_load_index = 0
        self._wm_thumb_width = 120
        self.cp_image_paths = []
        self.cp_image_items = {}
        self._cp_skeleton_cards = {}
        self._cp_load_index = 0
        self._cp_thumb_width = 120
        self.wm_cities = []
        self._load_cities()
        self._build_ui()
        # Loading overlay
        self.loading_overlay = LoadingOverlay(self.centralWidget(), "正在添加水印...")
        self.loading_overlay.hide()
        self.cp_loading_overlay = LoadingOverlay(self.centralWidget(), "正在压缩图片...")
        self.cp_loading_overlay.hide()
        self._generate_checkmark_icon()
        self.setStyleSheet(self._apply_qss())

    def resizeEvent(self, event):
        super().resizeEvent(event)
        cw = self.centralWidget()
        if cw and hasattr(self, 'loading_overlay'):
            self.loading_overlay.setGeometry(cw.rect())
        if cw and hasattr(self, 'cp_loading_overlay'):
            self.cp_loading_overlay.setGeometry(cw.rect())

    # ─── UI Construction ────────────────────────────────────────────────────────

    def _load_cities(self):
        if getattr(sys, 'frozen', False):
            base = sys._MEIPASS
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        json_path = os.path.join(base, "city_address.json")
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.wm_cities = list(data.keys())
        except Exception:
            self.wm_cities = []

    def _generate_checkmark_icon(self):
        """Create a white ✓ checkmark PNG from Heroicons SVG for QSS checkbox indicators."""
        # Heroicons check SVG (MIT license): render to QPixmap with white fill
        svg_data = (
            '<svg width="24" height="24" viewBox="0 0 24 24" fill="#ffffff" xmlns="http://www.w3.org/2000/svg">'
            '<path fill-rule="evenodd" clip-rule="evenodd" '
            'd="M19.916 4.62592C20.2607 4.85568 20.3538 5.32134 20.124 5.66598L11.124 19.166'
            'C10.9994 19.3529 10.7975 19.4742 10.5739 19.4963C10.3503 19.5184 10.1286 19.4392'
            ' 9.96967 19.2803L3.96967 13.2803C3.67678 12.9874 3.67678 12.5125 3.96967 12.2196'
            'C4.26256 11.9267 4.73744 11.9267 5.03033 12.2196L10.3834 17.5727L18.876 4.83393'
            'C19.1057 4.48929 19.5714 4.39616 19.916 4.62592Z" fill="#ffffff"/>'
            '</svg>'
        )
        # Render SVG to QPixmap
        from PySide6.QtSvg import QSvgRenderer
        renderer = QSvgRenderer(svg_data.encode('utf-8'))
        size = 18
        pixmap = QPixmap(size, size)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)
        renderer.render(painter)
        painter.end()
        # Save to temp directory
        tmp_dir = QStandardPaths.writableLocation(QStandardPaths.TempLocation)
        self._checkmark_icon_path = os.path.join(tmp_dir, "checkmark_white.png")
        pixmap.save(self._checkmark_icon_path, "PNG")

    def _apply_qss(self):
        """Inject checkmark icon path into QSS."""
        path = self._checkmark_icon_path.replace("/", "/")
        return MACOS_QSS.replace("__CHECKMARK_ICON__", path)

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # ── Body: Sidebar + Main ──
        body = QWidget()
        body_layout = QHBoxLayout(body)
        body_layout.setContentsMargins(0, 0, 0, 0)
        body_layout.setSpacing(0)

        # ── Sidebar ──
        sidebar = QFrame()
        sidebar.setObjectName("sidebar")
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(20, 0, 0, 0)
        sidebar_layout.setSpacing(0)
        sidebar_layout.addStretch(0)  # top padding via stretch before sections

        # EXCEL section
        excel_label = QLabel("EXCEL")
        excel_label.setObjectName("sidebar_label")
        sidebar_layout.addWidget(excel_label)

        self.sidebar_excel_btn = self._make_sidebar_item("📊", "Excel订单生成器", "sidebar_item_active")
        self.sidebar_excel_btn.clicked.connect(lambda: self._switch_page("excel"))
        sidebar_layout.addWidget(self.sidebar_excel_btn)

        # 图片 section
        img_label = QLabel("图片")
        img_label.setObjectName("sidebar_label")
        sidebar_layout.addWidget(img_label)

        self.sidebar_watermark_btn = self._make_sidebar_item("💧", "批量水印", "sidebar_item")
        self.sidebar_watermark_btn.clicked.connect(lambda: self._switch_page("watermark"))
        sidebar_layout.addWidget(self.sidebar_watermark_btn)

        self.sidebar_compress_btn = self._make_sidebar_item("📦", "图片压缩", "sidebar_item")
        self.sidebar_compress_btn.clicked.connect(lambda: self._switch_page("compress"))
        sidebar_layout.addWidget(self.sidebar_compress_btn)

        sidebar_layout.addStretch(1)

        body_layout.addWidget(sidebar)

        # ── Main Area ──
        self.page_stack = QStackedWidget()
        self.page_stack.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Build all pages
        self._build_excel_page()
        self._build_watermark_page()
        self._build_compress_page()

        body_layout.addWidget(self.page_stack, 1)

        root_layout.addWidget(body, 1)

    def _make_sidebar_item(self, icon_text, label_text, object_name):
        btn = QPushButton(f"{icon_text}  {label_text}")
        btn.setObjectName(object_name)
        btn.setFlat(True)
        btn.setCursor(Qt.PointingHandCursor)
        btn.setFixedHeight(30)
        return btn

    # ─── Excel Page ─────────────────────────────────────────────────────────────

    def _build_excel_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(16, 16, 20, 16)
        page_layout.setSpacing(0)

        # Scroll area for form card
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")

        card = QFrame()
        card.setObjectName("form_card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(18, 18, 22, 18)
        card_layout.setSpacing(0)

        # ── Excel名称 ──
        field_excel = self._make_field("📝", "Excel名称")
        self.excel_name_input = QLineEdit("订单表")
        self.excel_name_input.setObjectName("excel_name_input")
        self.excel_name_input.setPlaceholderText("请输入Excel文件名称")
        self.excel_name_input.textChanged.connect(self._on_excel_name_changed)
        field_excel.layout().addWidget(self.excel_name_input)
        card_layout.addWidget(field_excel)
        card_layout.addSpacing(12)

        # ── 选择已有Excel文件 ──
        field_file = self._make_field("📂", "选择已有Excel文件")
        file_row = QWidget()
        file_row_layout = QHBoxLayout(file_row)
        file_row_layout.setContentsMargins(0, 0, 0, 0)
        file_row_layout.setSpacing(10)
        file_row_layout.setAlignment(Qt.AlignVCenter)

        self.file_path_display = QLineEdit()
        self.file_path_display.setObjectName("filepath_display")
        self.file_path_display.setReadOnly(True)
        self.file_path_display.setPlaceholderText("未选择文件")
        file_row_layout.addWidget(self.file_path_display, 1)

        self.file_select_btn = QPushButton("选择文件")
        self.file_select_btn.setObjectName("file_select_btn")
        self.file_select_btn.setFixedSize(120, 35)
        self.file_select_btn.setCursor(Qt.PointingHandCursor)
        self.file_select_btn.clicked.connect(self._on_select_file)
        file_row_layout.addWidget(self.file_select_btn)

        self.file_clear_btn = QPushButton("清除")
        self.file_clear_btn.setObjectName("mac-btn-secondary")
        self.file_clear_btn.setMinimumWidth(60)
        self.file_clear_btn.setCursor(Qt.PointingHandCursor)
        self.file_clear_btn.clicked.connect(self._on_clear_file)
        self.file_clear_btn.setVisible(False)
        file_row_layout.addWidget(self.file_clear_btn)

        field_file.layout().addWidget(file_row)
        card_layout.addWidget(field_file)
        card_layout.addSpacing(12)

        # ── 月份(1-12) ──
        field_month = self._make_field("📅", "月份(1-12)")
        month_row = QWidget()
        month_row_layout = QHBoxLayout(month_row)
        month_row_layout.setContentsMargins(0, 0, 0, 0)
        month_row_layout.setSpacing(10)

        self.month_input = QLineEdit()
        self.month_input.setObjectName("month_input")
        self.month_input.setMaxLength(2)
        self.month_input.setPlaceholderText("请输入1-12")
        now_month = str(datetime.now().month)
        self.month_input.setText(now_month)
        month_row_layout.addWidget(self.month_input, 1)

        self.month_hint = QLabel(f"当前月份：{now_month}")
        self.month_hint.setObjectName("month_hint")
        self.month_hint.setMinimumWidth(80)
        month_row_layout.addWidget(self.month_hint)

        field_month.layout().addWidget(month_row)
        card_layout.addWidget(field_month)
        card_layout.addSpacing(12)

        # ── 3-column row: 销售日期 / 销售单号 / Pos单号 ──
        cols_widget = QWidget()
        cols_layout = QHBoxLayout(cols_widget)
        cols_layout.setContentsMargins(0, 0, 0, 0)
        cols_layout.setSpacing(14)

        # 销售日期总条数
        col_date = QWidget()
        col_date_layout = QVBoxLayout(col_date)
        col_date_layout.setContentsMargins(0, 0, 0, 0)
        col_date_layout.setSpacing(5)
        date_title = QLabel("销售日期总条数")
        date_title.setObjectName("field_title")
        col_date_layout.addWidget(date_title)
        self.date_count_input = QLineEdit()
        self.date_count_input.setObjectName("date_count_input")
        self.date_count_input.setPlaceholderText("请输入")
        col_date_layout.addWidget(self.date_count_input)
        cols_layout.addWidget(col_date, 1)

        # 销售单号总条数
        col_order = QWidget()
        col_order_layout = QVBoxLayout(col_order)
        col_order_layout.setContentsMargins(0, 0, 0, 0)
        col_order_layout.setSpacing(5)
        order_title = QLabel("销售单号总条数")
        order_title.setObjectName("field_title")
        col_order_layout.addWidget(order_title)
        self.order_count_input = QLineEdit()
        self.order_count_input.setObjectName("order_count_input")
        self.order_count_input.setPlaceholderText("请输入")
        col_order_layout.addWidget(self.order_count_input)
        cols_layout.addWidget(col_order, 1)

        # Pos单号总条数
        col_pos = QWidget()
        col_pos_layout = QVBoxLayout(col_pos)
        col_pos_layout.setContentsMargins(0, 0, 0, 0)
        col_pos_layout.setSpacing(5)
        pos_title = QLabel("Pos单号总条数")
        pos_title.setObjectName("field_title")
        col_pos_layout.addWidget(pos_title)
        self.pos_count_input = QLineEdit()
        self.pos_count_input.setObjectName("pos_count_input")
        self.pos_count_input.setPlaceholderText("请输入")
        col_pos_layout.addWidget(self.pos_count_input)
        cols_layout.addWidget(col_pos, 1)

        card_layout.addWidget(cols_widget)
        card_layout.addSpacing(12)

        # ── 开始生成按钮 ──
        self.start_btn = QPushButton("开始生成")
        self.start_btn.setObjectName("generate_btn")
        self.start_btn.setCursor(Qt.PointingHandCursor)
        self.start_btn.clicked.connect(self.start_generation)
        card_layout.addWidget(self.start_btn)
        card_layout.addSpacing(12)

        # ── 操作日志 ──
        log_header = QWidget()
        log_header_layout = QHBoxLayout(log_header)
        log_header_layout.setContentsMargins(0, 0, 0, 0)
        log_header_layout.setSpacing(0)

        log_title = QLabel("📋  操作日志")
        log_title.setObjectName("log_title")
        log_header_layout.addWidget(log_title)
        log_header_layout.addStretch()

        self.copy_log_btn = QPushButton("复制日志")
        self.copy_log_btn.setObjectName("copy_log_btn")
        self.copy_log_btn.setCursor(Qt.PointingHandCursor)
        self.copy_log_btn.clicked.connect(self._on_copy_log)
        log_header_layout.addWidget(self.copy_log_btn)

        card_layout.addWidget(log_header)
        card_layout.addSpacing(8)

        self.log_area = QTextEdit()
        self.log_area.setObjectName("log_area")
        self.log_area.setReadOnly(True)
        self.log_area.setMinimumHeight(175)
        card_layout.addWidget(self.log_area, 1)

        scroll.setWidget(card)
        page_layout.addWidget(scroll, 1)

        self.page_stack.addWidget(page)

    def _make_field(self, icon, title_text):
        """Create a field container with icon+title label."""
        field = QWidget()
        layout = QVBoxLayout(field)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)
        title = QLabel(f"{icon}  {title_text}")
        title.setObjectName("field_title")
        layout.addWidget(title)
        field.setLayout(layout)
        return field

    # ─── Watermark Page ──────────────────────────────────────────────────────────

    def _build_watermark_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(16, 16, 20, 16)
        page_layout.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")

        card = QFrame()
        card.setObjectName("form_card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(18, 18, 22, 18)
        card_layout.setSpacing(0)

        # ── Section 1: Image Display Box ──
        img_card = QFrame()
        img_card.setObjectName("wm_image_card")
        img_layout = QVBoxLayout(img_card)
        img_layout.setContentsMargins(0, 0, 0, 0)
        img_layout.setSpacing(10)

        # Title row
        title_row = QWidget()
        title_layout = QHBoxLayout(title_row)
        title_layout.setContentsMargins(0, 0, 0, 0)
        title_label = QLabel("🖼️ 图片显示框")
        title_label.setObjectName("wm_image_title")
        title_layout.addWidget(title_label)
        title_layout.addStretch()

        self.wm_select_btn = QPushButton("选择文件")
        self.wm_select_btn.setObjectName("wm_select_btn")
        self.wm_select_btn.setCursor(Qt.PointingHandCursor)
        self.wm_select_btn.clicked.connect(self._on_wm_select_images)
        title_layout.addWidget(self.wm_select_btn)

        self.wm_clear_btn = QPushButton("清空")
        self.wm_clear_btn.setObjectName("wm_clear_btn")
        self.wm_clear_btn.setCursor(Qt.PointingHandCursor)
        self.wm_clear_btn.clicked.connect(self._on_wm_clear_images)
        self.wm_clear_btn.setVisible(False)
        title_layout.addWidget(self.wm_clear_btn)

        img_layout.addWidget(title_row)

        # Empty state — original height when no images
        self.wm_empty_state = QWidget()
        self.wm_empty_state.setMinimumHeight(160)
        self.wm_empty_state.setMaximumHeight(160)
        empty_layout = QVBoxLayout(self.wm_empty_state)
        empty_layout.setAlignment(Qt.AlignCenter)
        empty_icon = QLabel("🖼️")
        empty_icon.setObjectName("wm_empty_icon")
        empty_icon.setAlignment(Qt.AlignCenter)
        empty_layout.addWidget(empty_icon)
        empty_text = QLabel("点击「选择文件」添加图片")
        empty_text.setObjectName("wm_empty_text")
        empty_text.setAlignment(Qt.AlignCenter)
        empty_layout.addWidget(empty_text)
        empty_hint = QLabel("支持 JPG / PNG / WebP 格式")
        empty_hint.setObjectName("wm_empty_hint")
        empty_hint.setAlignment(Qt.AlignCenter)
        empty_layout.addWidget(empty_hint)
        img_layout.addWidget(self.wm_empty_state)

        # Scroll area — height adjusted dynamically based on row count
        self.wm_image_scroll = QScrollArea()
        self.wm_image_scroll.setObjectName("wm_image_scroll")
        self.wm_image_scroll.setWidgetResizable(True)
        self.wm_image_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.wm_image_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.wm_image_scroll.setVisible(False)

        self.wm_image_grid_container = QWidget()
        grid_outer_layout = QVBoxLayout(self.wm_image_grid_container)
        grid_outer_layout.setContentsMargins(0, 0, 0, 0)
        self.wm_image_grid_layout = QGridLayout()
        self.wm_image_grid_layout.setSpacing(10)
        grid_outer_layout.addLayout(self.wm_image_grid_layout)

        self.wm_image_scroll.setWidget(self.wm_image_grid_container)
        img_layout.addWidget(self.wm_image_scroll)

        card_layout.addWidget(img_card)
        card_layout.addSpacing(12)

        # ── Section 2: City Filter ──
        city_card = QFrame()
        city_card.setObjectName("wm_city_card")
        city_layout = QVBoxLayout(city_card)
        city_layout.setContentsMargins(0, 0, 0, 0)
        city_layout.setSpacing(10)

        city_title = QLabel("📍 城市选择")
        city_title.setObjectName("wm_city_title")
        city_layout.addWidget(city_title)

        city_grid = QWidget()
        city_grid.setObjectName("wm_city_grid")
        city_grid_layout = QGridLayout()
        city_grid_layout.setSpacing(14)
        city_grid_layout.setContentsMargins(12, 12, 12, 12)

        # "全部城市" checkbox
        self.wm_all_city_checkbox = QCheckBox("全部城市")
        self.wm_all_city_checkbox.setObjectName("wm_all_city_checkbox")
        self.wm_all_city_checkbox.setChecked(True)
        self.wm_all_city_checkbox.stateChanged.connect(self._on_wm_all_cities_toggled)
        city_grid_layout.addWidget(self.wm_all_city_checkbox, 0, 0)

        # Individual city checkboxes
        self.wm_city_checkboxes = []
        cols = 5
        for i, city in enumerate(self.wm_cities):
            cb = QCheckBox(city)
            cb.setObjectName("wm_city_checkbox")
            cb.setChecked(True)
            cb.stateChanged.connect(self._on_wm_city_checkbox_changed)
            row = (i + 1) // cols
            col = (i + 1) % cols
            city_grid_layout.addWidget(cb, row, col)
            self.wm_city_checkboxes.append(cb)

        city_grid.setLayout(city_grid_layout)
        city_layout.addWidget(city_grid)

        card_layout.addWidget(city_card)
        card_layout.addSpacing(12)

        # ── Section 3: Date Selector ──
        date_card = QFrame()
        date_card.setObjectName("wm_date_card")
        date_layout = QVBoxLayout(date_card)
        date_layout.setContentsMargins(0, 0, 0, 0)
        date_layout.setSpacing(10)

        date_title = QLabel("📅 日期选择")
        date_title.setObjectName("wm_date_title")
        date_layout.addWidget(date_title)

        date_row = QWidget()
        date_row_layout = QHBoxLayout(date_row)
        date_row_layout.setContentsMargins(0, 0, 0, 0)
        date_row_layout.setSpacing(10)

        year_label = QLabel("年份")
        year_label.setObjectName("wm_date_label")
        date_row_layout.addWidget(year_label)

        self.wm_year_input = QLineEdit(str(datetime.now().year))
        self.wm_year_input.setObjectName("wm_year_input")
        self.wm_year_input.setFixedWidth(100)
        self.wm_year_input.setPlaceholderText("年份")
        date_row_layout.addWidget(self.wm_year_input)

        month_label = QLabel("月份")
        month_label.setObjectName("wm_date_label")
        date_row_layout.addWidget(month_label)

        self.wm_month_input = QLineEdit(str(datetime.now().month))
        self.wm_month_input.setObjectName("wm_month_input")
        self.wm_month_input.setFixedWidth(100)
        self.wm_month_input.setPlaceholderText("月份(1-12)")
        date_row_layout.addWidget(self.wm_month_input)

        date_row_layout.addStretch()
        date_layout.addWidget(date_row)

        card_layout.addWidget(date_card)
        card_layout.addSpacing(16)

        # ── Section 4: Start Button ──
        start_row = QWidget()
        start_row_layout = QHBoxLayout(start_row)
        start_row_layout.setContentsMargins(0, 0, 0, 0)
        start_row_layout.addStretch()

        self.wm_start_btn = QPushButton("▶ 开始添加水印")
        self.wm_start_btn.setObjectName("wm_start_btn")
        self.wm_start_btn.setCursor(Qt.PointingHandCursor)
        self.wm_start_btn.clicked.connect(self._on_wm_start)
        start_row_layout.addWidget(self.wm_start_btn)
        start_row_layout.addStretch()
        card_layout.addWidget(start_row)

        # ── Section 5: Watermark Operation Log ──
        wm_log_header = QWidget()
        wm_log_header_layout = QHBoxLayout(wm_log_header)
        wm_log_header_layout.setContentsMargins(0, 0, 0, 0)
        wm_log_header_layout.setSpacing(0)

        wm_log_title = QLabel("📋  操作日志")
        wm_log_title.setObjectName("log_title")
        wm_log_header_layout.addWidget(wm_log_title)
        wm_log_header_layout.addStretch()

        self.wm_copy_log_btn = QPushButton("复制日志")
        self.wm_copy_log_btn.setObjectName("copy_log_btn")
        self.wm_copy_log_btn.setCursor(Qt.PointingHandCursor)
        self.wm_copy_log_btn.clicked.connect(self._on_wm_copy_log)
        wm_log_header_layout.addWidget(self.wm_copy_log_btn)

        card_layout.addWidget(wm_log_header)
        card_layout.addSpacing(8)

        self.wm_log_area = QTextEdit()
        self.wm_log_area.setObjectName("log_area")
        self.wm_log_area.setReadOnly(True)
        self.wm_log_area.setMinimumHeight(150)
        card_layout.addWidget(self.wm_log_area, 1)

        scroll.setWidget(card)
        page_layout.addWidget(scroll, 1)

        self.page_stack.addWidget(page)
        self._watermark_page_idx = self.page_stack.count() - 1

    # ─── Compress Page ──────────────────────────────────────────────────────────────

    def _build_compress_page(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(16, 16, 20, 16)
        page_layout.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")

        card = QFrame()
        card.setObjectName("form_card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(18, 18, 22, 18)
        card_layout.setSpacing(0)

        # ── 图片显示框 ──
        img_card = QFrame()
        img_card.setObjectName("cp_image_card")
        img_layout = QVBoxLayout(img_card)
        img_layout.setContentsMargins(0, 0, 0, 0)
        img_layout.setSpacing(8)

        # Title row
        title_row = QHBoxLayout()
        title_label = QLabel("📦 图片显示框")
        title_label.setObjectName("cp_image_title")
        title_row.addWidget(title_label)
        title_row.addStretch()
        self.cp_select_btn = QPushButton("选择文件")
        self.cp_select_btn.setObjectName("cp_select_btn")
        self.cp_select_btn.setCursor(Qt.PointingHandCursor)
        self.cp_select_btn.clicked.connect(self._on_cp_select_images)
        title_row.addWidget(self.cp_select_btn)
        self.cp_clear_btn = QPushButton("清空")
        self.cp_clear_btn.setObjectName("cp_clear_btn")
        self.cp_clear_btn.setCursor(Qt.PointingHandCursor)
        self.cp_clear_btn.clicked.connect(self._on_cp_clear_images)
        title_row.addWidget(self.cp_clear_btn)
        img_layout.addLayout(title_row)

        # Empty state
        self.cp_empty_state = QWidget()
        cp_empty_layout = QVBoxLayout(self.cp_empty_state)
        cp_empty_layout.setAlignment(Qt.AlignCenter)
        cp_empty_icon = QLabel("📦")
        cp_empty_icon.setObjectName("cp_empty_icon")
        cp_empty_icon.setAlignment(Qt.AlignCenter)
        cp_empty_layout.addWidget(cp_empty_icon)
        cp_empty_text = QLabel("点击「选择文件」添加图片")
        cp_empty_text.setObjectName("cp_empty_text")
        cp_empty_text.setAlignment(Qt.AlignCenter)
        cp_empty_layout.addWidget(cp_empty_text)
        cp_empty_hint = QLabel("支持 JPG / PNG / WebP 格式")
        cp_empty_hint.setObjectName("cp_empty_hint")
        cp_empty_hint.setAlignment(Qt.AlignCenter)
        cp_empty_layout.addWidget(cp_empty_hint)
        img_layout.addWidget(self.cp_empty_state)

        # Image grid in scroll area
        self.cp_image_scroll = QScrollArea()
        self.cp_image_scroll.setObjectName("cp_image_scroll")
        self.cp_image_scroll.setWidgetResizable(True)
        self.cp_image_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.cp_image_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.cp_image_scroll.setFrameShape(QFrame.NoFrame)
        self.cp_image_scroll.hide()
        cp_grid_container = QWidget()
        self.cp_image_grid_layout = QGridLayout(cp_grid_container)
        self.cp_image_grid_layout.setSpacing(10)
        self.cp_image_grid_layout.setContentsMargins(0, 0, 0, 0)
        self.cp_image_scroll.setWidget(cp_grid_container)
        img_layout.addWidget(self.cp_image_scroll)

        self.cp_clear_btn.hide()
        card_layout.addWidget(img_card)
        card_layout.addSpacing(12)

        # ── 压缩方式选择 ──
        method_card = QFrame()
        method_card.setObjectName("cp_method_card")
        method_layout = QVBoxLayout(method_card)
        method_layout.setContentsMargins(16, 16, 16, 16)
        method_layout.setSpacing(8)

        method_title = QLabel("⚙ 压缩方式")
        method_title.setObjectName("cp_method_title")
        method_layout.addWidget(method_title)

        self.cp_method_group = QButtonGroup(self)

        # 无损压缩
        lossless_row = QVBoxLayout()
        lossless_row.setSpacing(2)
        self.cp_radio_lossless = QRadioButton("无损压缩")
        self.cp_radio_lossless.setObjectName("cp_method_radio")
        self.cp_radio_lossless.setChecked(True)
        self.cp_method_group.addButton(self.cp_radio_lossless, 0)
        lossless_row.addWidget(self.cp_radio_lossless)
        lossless_note = QLabel("保持原格式，最高质量参数重编码优化。预估压缩10~30%体积")
        lossless_note.setObjectName("cp_method_note")
        lossless_row.addWidget(lossless_note)
        method_layout.addLayout(lossless_row)

        # 有损压缩
        lossy_row = QVBoxLayout()
        lossy_row.setSpacing(2)
        self.cp_radio_lossy = QRadioButton("有损压缩")
        self.cp_radio_lossy.setObjectName("cp_method_radio")
        self.cp_method_group.addButton(self.cp_radio_lossy, 1)
        lossy_row.addWidget(self.cp_radio_lossy)
        lossy_note = QLabel("降低画质换取更小体积，肉眼几乎看不出差异。预估压缩40~60%体积")
        lossy_note.setObjectName("cp_method_note")
        lossy_row.addWidget(lossy_note)
        method_layout.addLayout(lossy_row)

        self.cp_method_group.buttonClicked.connect(self._cp_on_method_changed)

        card_layout.addWidget(method_card)
        card_layout.addSpacing(12)

        # ── 压缩预估统计 ──
        stats_card = QFrame()
        stats_card.setObjectName("cp_stats_card")
        stats_layout = QVBoxLayout(stats_card)
        stats_layout.setContentsMargins(0, 0, 0, 0)
        stats_layout.setSpacing(10)

        stats_title = QLabel("📊 压缩预估")
        stats_title.setObjectName("cp_stats_title")
        stats_layout.addWidget(stats_title)

        stats_grid = QGridLayout()
        stats_grid.setSpacing(10)

        # Stat item 1: 当前占用内存
        stat1 = QFrame()
        stat1.setObjectName("cp_stat_item")
        stat1_layout = QVBoxLayout(stat1)
        stat1_layout.setSpacing(4)
        self.cp_stat_label1 = QLabel("当前占用内存")
        self.cp_stat_label1.setObjectName("cp_stat_label")
        stat1_layout.addWidget(self.cp_stat_label1)
        self.cp_stat_value1 = QLabel("0 MB")
        self.cp_stat_value1.setObjectName("cp_stat_value_orange")
        stat1_layout.addWidget(self.cp_stat_value1)
        self.cp_stat_sub1 = QLabel("0 张图片 · 未压缩")
        self.cp_stat_sub1.setObjectName("cp_stat_sub")
        stat1_layout.addWidget(self.cp_stat_sub1)
        stats_grid.addWidget(stat1, 0, 0)

        # Stat item 2: 压缩后预估
        stat2 = QFrame()
        stat2.setObjectName("cp_stat_item")
        stat2_layout = QVBoxLayout(stat2)
        stat2_layout.setSpacing(4)
        self.cp_stat_label2 = QLabel("压缩后预估")
        self.cp_stat_label2.setObjectName("cp_stat_label")
        stat2_layout.addWidget(self.cp_stat_label2)
        self.cp_stat_value2 = QLabel("0 MB")
        self.cp_stat_value2.setObjectName("cp_stat_value_green")
        stat2_layout.addWidget(self.cp_stat_value2)
        self.cp_stat_sub2 = QLabel("压缩率约 70%")
        self.cp_stat_sub2.setObjectName("cp_stat_sub")
        stat2_layout.addWidget(self.cp_stat_sub2)
        stats_grid.addWidget(stat2, 0, 1)

        stats_layout.addLayout(stats_grid)
        card_layout.addWidget(stats_card)
        card_layout.addSpacing(12)

        # ── 开始压缩按钮 ──
        start_btn_row = QHBoxLayout()
        start_btn_row.setAlignment(Qt.AlignCenter)
        self.cp_start_btn = QPushButton("▶ 开始压缩")
        self.cp_start_btn.setObjectName("cp_start_btn")
        self.cp_start_btn.setFixedSize(220, 40)
        self.cp_start_btn.setCursor(Qt.PointingHandCursor)
        self.cp_start_btn.clicked.connect(self._on_cp_start)
        start_btn_row.addWidget(self.cp_start_btn)
        card_layout.addLayout(start_btn_row)
        card_layout.addSpacing(12)

        # ── 操作日志 ──
        cp_log_header = QWidget()
        cp_log_header_layout = QHBoxLayout(cp_log_header)
        cp_log_header_layout.setContentsMargins(0, 0, 0, 0)
        cp_log_header_layout.setSpacing(0)

        cp_log_title = QLabel("📋  操作日志")
        cp_log_title.setObjectName("log_title")
        cp_log_header_layout.addWidget(cp_log_title)
        cp_log_header_layout.addStretch()

        self.cp_copy_log_btn = QPushButton("复制日志")
        self.cp_copy_log_btn.setObjectName("copy_log_btn")
        self.cp_copy_log_btn.setCursor(Qt.PointingHandCursor)
        self.cp_copy_log_btn.clicked.connect(self._on_cp_copy_log)
        cp_log_header_layout.addWidget(self.cp_copy_log_btn)

        card_layout.addWidget(cp_log_header)
        card_layout.addSpacing(8)

        self.cp_log_area = QTextEdit()
        self.cp_log_area.setObjectName("log_area")
        self.cp_log_area.setReadOnly(True)
        self.cp_log_area.setMinimumHeight(150)
        card_layout.addWidget(self.cp_log_area, 1)

        scroll.setWidget(card)
        page_layout.addWidget(scroll, 1)
        self.page_stack.addWidget(page)
        self._compress_page_idx = self.page_stack.count() - 1

    def _build_placeholder_page(self, icon, text, page_id):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(16, 16, 20, 16)
        page_layout.setAlignment(Qt.AlignCenter)

        icon_label = QLabel(icon)
        icon_label.setObjectName("placeholder_icon")
        icon_label.setAlignment(Qt.AlignCenter)
        page_layout.addWidget(icon_label)

        text_label = QLabel(text)
        text_label.setObjectName("placeholder_text")
        text_label.setAlignment(Qt.AlignCenter)
        page_layout.addWidget(text_label)

        self.page_stack.addWidget(page)
        self._placeholder_pages[page_id] = self.page_stack.count() - 1

    # ─── Sidebar Navigation ─────────────────────────────────────────────────────

    _placeholder_pages = {}

    def _switch_page(self, page_id):
        items = {
            "excel": self.sidebar_excel_btn,
            "watermark": self.sidebar_watermark_btn,
            "compress": self.sidebar_compress_btn,
        }
        for key, btn in items.items():
            if key == page_id:
                btn.setObjectName("sidebar_item_active")
            else:
                btn.setObjectName("sidebar_item")
        # Force style refresh
        for btn in items.values():
            btn.style().unpolish(btn)
            btn.style().polish(btn)

        page_indices = {
            "excel": 0,
            "watermark": self._watermark_page_idx,
            "compress": self._compress_page_idx,
        }
        idx = page_indices.get(page_id)
        if idx is not None:
            self.page_stack.setCurrentIndex(idx)

    # ─── Interaction Logic ──────────────────────────────────────────────────────

    # ── Compress Interaction ──

    def _on_cp_select_images(self):
        dlg = ImagePickerDialog(self, set(self.cp_image_paths))
        if dlg.exec() == QDialog.Accepted:
            new_paths = dlg.get_selected_paths()
            for p in new_paths:
                if p not in self.cp_image_paths:
                    self.cp_image_paths.append(p)
            self._cp_refresh_grid()

    def _on_cp_clear_images(self):
        self.cp_image_paths.clear()
        self._cp_refresh_grid()

    def _on_cp_delete_image(self, path):
        if path in self.cp_image_paths:
            self.cp_image_paths.remove(path)
        self._cp_refresh_grid()

    def _on_cp_preview_image(self, path):
        pixmap = self._load_rotated_pixmap(path, target_width=800)
        if pixmap.isNull():
            QMessageBox.warning(self, "提示", "无法加载图片")
            return
        h_ratio = 600 / pixmap.height() if pixmap.height() > 600 else 1
        w_ratio = 800 / pixmap.width() if pixmap.width() > 800 else 1
        scale = min(h_ratio, w_ratio)
        if scale < 1:
            pixmap = pixmap.scaled(int(pixmap.width() * scale), int(pixmap.height() * scale), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        dlg = QDialog(self)
        dlg.setWindowTitle(os.path.basename(path))
        dlg.setObjectName("cp_preview_dialog")
        dlg.setStyleSheet("QDialog#cp_preview_dialog { background: #1d1d1f; }")
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(8, 8, 8, 8)
        label = QLabel()
        label.setPixmap(pixmap)
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)
        dlg.exec()

    def _cp_refresh_grid(self):
        # Clear old items (both real and skeleton)
        for path, item in self.cp_image_items.items():
            item.deleteLater()
        self.cp_image_items.clear()
        for item in self._cp_skeleton_cards.values():
            item.deleteLater()
        self._cp_skeleton_cards.clear()
        self._cp_load_index = 0

        # Toggle empty state vs grid
        has_images = len(self.cp_image_paths) > 0
        self.cp_empty_state.setVisible(not has_images)
        self.cp_image_scroll.setVisible(has_images)
        self.cp_clear_btn.setVisible(has_images)

        if not has_images:
            self._cp_update_stats()
            return

        # Calculate thumbnail width
        scroll_width = self.cp_image_scroll.width()
        if scroll_width < 100:
            scroll_width = self.width() - 50
        spacing_total = 10 * 3
        padding_total = 8 * 4
        thumb_width = max(80, (scroll_width - spacing_total - padding_total) // 4)
        self._cp_thumb_width = thumb_width

        # Create skeleton cards for all images immediately
        for i, path in enumerate(self.cp_image_paths):
            row, col = i // 4, i % 4
            skeleton = self._cp_create_skeleton_card(thumb_width)
            self._cp_skeleton_cards[i] = skeleton
            self.cp_image_grid_layout.addWidget(skeleton, row, col)

        # Set initial scroll height
        ROW_HEIGHT = 250
        GRID_SPACING = 10
        PEEK_HEIGHT = ROW_HEIGHT // 4
        total_rows = (len(self.cp_image_paths) + 3) // 4
        if total_rows <= 2:
            scroll_h = total_rows * ROW_HEIGHT + (total_rows - 1) * GRID_SPACING
        else:
            scroll_h = 2 * ROW_HEIGHT + GRID_SPACING + PEEK_HEIGHT
        self.cp_image_scroll.setMinimumHeight(scroll_h)
        self.cp_image_scroll.setMaximumHeight(scroll_h)

        self._cp_update_stats()
        # Start replacing skeletons with real images one by one
        QTimer.singleShot(50, self._cp_load_next_image)

    def _cp_create_skeleton_card(self, tw):
        card = QFrame()
        card.setObjectName("cp_skeleton")
        card.setMinimumHeight(250)
        card.setMaximumWidth(tw + 8)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(4, 4, 4, 4)
        card_layout.setSpacing(2)
        # Fake delete row
        top_row = QWidget()
        top_layout = QHBoxLayout(top_row)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.addStretch()
        card_layout.addWidget(top_row)
        # Shimmer placeholder
        shimmer = QLabel()
        shimmer.setObjectName("cp_skeleton_shimmer")
        shimmer.setFixedSize(tw, 120)
        card_layout.addWidget(shimmer, 1)
        # Fake filename bar
        card_layout.addWidget(QLabel())
        return card

    def _cp_load_next_image(self):
        if self._cp_load_index >= len(self.cp_image_paths):
            return

        path = self.cp_image_paths[self._cp_load_index]
        index = self._cp_load_index
        row = index // 4
        col = index % 4
        tw = self._cp_thumb_width

        # Remove skeleton card at this position
        skeleton = self._cp_skeleton_cards.get(index)
        if skeleton:
            self.cp_image_grid_layout.removeWidget(skeleton)
            skeleton.deleteLater()
            del self._cp_skeleton_cards[index]

        # Create real card
        item = QFrame()
        item.setObjectName("cp_image_item")
        item.setMinimumHeight(250)
        item.setMaximumWidth(tw + 8)
        item_layout = QVBoxLayout(item)
        item_layout.setContentsMargins(4, 4, 4, 4)
        item_layout.setSpacing(2)

        # Top row: delete button at right
        top_row = QWidget()
        top_layout = QHBoxLayout(top_row)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.addStretch()

        del_btn = QPushButton("✕")
        del_btn.setObjectName("cp_delete_btn")
        del_btn.setCursor(Qt.PointingHandCursor)
        del_btn.clicked.connect(lambda checked=False, p=path: self._on_cp_delete_image(p))
        top_layout.addWidget(del_btn)
        item_layout.addWidget(top_row)

        # Thumbnail (clickable for preview)
        pixmap = self._load_rotated_pixmap(path, target_width=tw)
        thumb_btn = QPushButton()
        thumb_btn.setObjectName("cp_img_btn")
        thumb_btn.setCursor(Qt.PointingHandCursor)
        if not pixmap.isNull():
            scaled = pixmap.scaled(tw, 300, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            thumb_btn.setIcon(QIcon(scaled))
            thumb_btn.setIconSize(scaled.size())
            thumb_btn.setFixedSize(scaled.size())
        else:
            thumb_btn.setText("无法加载")
            thumb_btn.setFixedSize(tw, 200)
        thumb_btn.clicked.connect(lambda checked=False, p=path: self._on_cp_preview_image(p))
        item_layout.addWidget(thumb_btn, 1)

        # Filename label
        name_label = QLabel(os.path.basename(path))
        name_label.setObjectName("cp_filename_label")
        name_label.setAlignment(Qt.AlignCenter)
        item_layout.addWidget(name_label)

        self.cp_image_grid_layout.addWidget(item, row, col)
        self.cp_image_items[path] = item

        self._cp_load_index += 1
        self._cp_update_stats()
        QTimer.singleShot(0, self._cp_load_next_image)

    def _cp_update_stats(self):
        total_bytes = 0
        count = len(self.cp_image_paths)
        estimated_bytes = 0
        is_lossless = self.cp_method_group.button(0).isChecked()

        for path in self.cp_image_paths:
            try:
                orig_size = os.path.getsize(path)
                total_bytes += orig_size

                # Estimate compressed size by doing a lightweight in-memory compress
                img = Image.open(path)
                img = ImageOps.exif_transpose(img)
                ext = os.path.splitext(path)[1].lower()

                from io import BytesIO
                buf = BytesIO()

                if is_lossless:
                    if ext in ('.jpg', '.jpeg'):
                        img.convert('RGB').save(buf, 'JPEG', quality=85, optimize=True, progressive=True)
                    elif ext == '.png':
                        img.save(buf, 'PNG', compress_level=9, optimize=True)
                    elif ext == '.webp':
                        img.save(buf, 'WEBP', lossless=True, method=6)
                    else:
                        img.convert('RGB').save(buf, 'PNG', compress_level=9, optimize=True)
                else:
                    if ext in ('.jpg', '.jpeg'):
                        img.convert('RGB').save(buf, 'JPEG', quality=75, optimize=True, progressive=True)
                    elif ext == '.png':
                        img.convert('RGB').save(buf, 'JPEG', quality=75, optimize=True, progressive=True)
                    elif ext == '.webp':
                        img.convert('RGB').save(buf, 'WEBP', quality=75, method=4)
                    else:
                        img.convert('RGB').save(buf, 'JPEG', quality=75, optimize=True, progressive=True)

                estimated_bytes += buf.tell()
                buf.close()
            except Exception:
                # Fallback: use rough percentage estimate
                try:
                    orig_size = os.path.getsize(path)
                    total_bytes += orig_size
                    if is_lossless:
                        estimated_bytes += orig_size * 0.70
                    else:
                        estimated_bytes += orig_size * 0.50
                except OSError:
                    pass

        total_mb = total_bytes / 1024 / 1024
        estimated_mb = estimated_bytes / 1024 / 1024

        if is_lossless:
            ratio = (1 - estimated_bytes / total_bytes) * 100 if total_bytes > 0 else 0
        else:
            ratio = (1 - estimated_bytes / total_bytes) * 100 if total_bytes > 0 else 0

        if total_mb >= 1:
            self.cp_stat_value1.setText(f"{total_mb:.1f} MB")
        elif total_bytes > 0:
            self.cp_stat_value1.setText(f"{total_bytes / 1024:.0f} KB")
        else:
            self.cp_stat_value1.setText("0 MB")

        if estimated_mb >= 1:
            self.cp_stat_value2.setText(f"{estimated_mb:.1f} MB")
        elif estimated_bytes > 0:
            self.cp_stat_value2.setText(f"{estimated_bytes / 1024:.0f} KB")
        else:
            self.cp_stat_value2.setText("0 MB")

        self.cp_stat_sub1.setText(f"{count} 张图片 · 未压缩")
        if count > 0 and total_bytes > 0:
            self.cp_stat_sub2.setText(f"预估压缩 {ratio:.0f}% 体积")
        else:
            self.cp_stat_sub2.setText("压缩率约 —")

    def _cp_on_method_changed(self):
        self._cp_update_stats()

    def _cp_show_loading(self):
        self.cp_loading_overlay.show_loading()
        QApplication.processEvents()

    def _cp_hide_loading(self):
        self.cp_loading_overlay.hide()

    def _cp_show_success(self, msg):
        self.cp_loading_overlay.show_success(msg)

    def _on_cp_start(self):
        try:
            if not self.cp_image_paths:
                QMessageBox.warning(self, "提示", "请先选择图片")
                return

            self.cp_log_area.clear()

            is_lossless = self.cp_method_group.button(0).isChecked()
            method_name = "无损压缩" if is_lossless else "有损压缩"
            desktop = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)

            compressed_dir = os.path.join(desktop, "已压缩图片")
            src_from_compressed = any(
                os.path.abspath(p).startswith(os.path.abspath(compressed_dir))
                for p in self.cp_image_paths
            )

            if src_from_compressed:
                output_dir = os.path.join(desktop, "已压缩图片-新")
            else:
                output_dir = compressed_dir

            os.makedirs(output_dir, exist_ok=True)

            self._cp_log_message("压缩任务启动...")
            self._cp_log_message(f"压缩方式: {method_name}")
            self._cp_log_message(f"图片数量: {len(self.cp_image_paths)} 张")
            self._cp_log_message(f"输出目录: {output_dir}")

            self._cp_show_loading()
            success_count = 0
            fail_list = []

            for idx, path in enumerate(self.cp_image_paths):
                name = os.path.basename(path)
                self._cp_log_message(f"[{idx+1}/{len(self.cp_image_paths)}] 正在压缩: {name}")
                QApplication.processEvents()
                try:
                    img = Image.open(path)
                    img = ImageOps.exif_transpose(img)
                    ext = os.path.splitext(name)[1].lower()
                    orig_size = os.path.getsize(path)

                    if is_lossless:
                        if ext in ('.jpg', '.jpeg'):
                            # JPEG无损: 保持JPEG格式，最高质量+优化参数
                            img_save = img.convert('RGB')
                            out_path = os.path.join(output_dir, name)
                            img_save.save(out_path, 'JPEG', quality=85, optimize=True, progressive=True)
                            self._cp_log_message(f"  JPEG无损优化 (quality=85, optimize)")
                        elif ext == '.png':
                            # PNG无损: 最高压缩等级+优化
                            out_path = os.path.join(output_dir, name)
                            img.save(out_path, 'PNG', compress_level=9, optimize=True)
                            self._cp_log_message(f"  PNG无损优化 (compress_level=9)")
                        elif ext == '.webp':
                            # WebP无损: 最高压缩方法
                            out_path = os.path.join(output_dir, name)
                            img.save(out_path, 'WEBP', lossless=True, method=6)
                            self._cp_log_message(f"  WebP无损优化 (lossless, method=6)")
                        else:
                            # 其他格式→PNG无损
                            img_save = img.convert('RGB')
                            out_name = os.path.splitext(name)[0] + '.png'
                            out_path = os.path.join(output_dir, out_name)
                            img_save.save(out_path, 'PNG', compress_level=9, optimize=True)
                            self._cp_log_message(f"  其他格式 → PNG无损")
                    else:
                        if ext in ('.jpg', '.jpeg'):
                            img_save = img.convert('RGB')
                            out_path = os.path.join(output_dir, name)
                            img_save.save(out_path, 'JPEG', quality=75, optimize=True, progressive=True)
                            self._cp_log_message(f"  JPEG有损 (quality=75)")
                        elif ext == '.png':
                            img_save = img.convert('RGB')
                            out_name = os.path.splitext(name)[0] + '.jpg'
                            out_path = os.path.join(output_dir, out_name)
                            img_save.save(out_path, 'JPEG', quality=75, optimize=True, progressive=True)
                            self._cp_log_message(f"  PNG → JPEG有损 (quality=75)")
                        elif ext == '.webp':
                            img_save = img.convert('RGB')
                            out_path = os.path.join(output_dir, name)
                            img_save.save(out_path, 'WEBP', quality=75, method=4)
                            self._cp_log_message(f"  WebP有损 (quality=75)")
                        else:
                            img_save = img.convert('RGB')
                            out_name = os.path.splitext(name)[0] + '.jpg'
                            out_path = os.path.join(output_dir, out_name)
                            img_save.save(out_path, 'JPEG', quality=75, optimize=True, progressive=True)
                            self._cp_log_message(f"  其他格式 → JPEG有损 (quality=75)")

                    new_size = os.path.getsize(out_path)
                    saved_pct = (1 - new_size / orig_size) * 100 if orig_size > 0 else 0
                    self._cp_log_message(f"  原始: {orig_size/1024/1024:.1f}MB → 压缩后: {new_size/1024/1024:.1f}MB (节省{saved_pct:.0f}%)")
                    self._cp_log_message(f"  已保存: {out_path}")
                    success_count += 1
                except Exception as e:
                    fail_msg = f"{name}: {str(e)}"
                    fail_list.append(fail_msg)
                    self._cp_log_message(f"压缩失败 {name}: {e}", is_error=True)
                    self._cp_log_message(traceback.format_exc(), is_error=True)
                QApplication.processEvents()

            self._cp_hide_loading()

            if fail_list:
                self._cp_log_message(f"压缩完成！成功 {success_count} 张, 失败 {len(fail_list)} 张", is_error=True)
                msg = f"压缩完成！成功 {success_count} 张\n失败 {len(fail_list)} 张:\n" + "\n".join(fail_list)
                QMessageBox.warning(self, "压缩结果", msg)
            else:
                self._cp_log_message(f"全部完成！共 {success_count} 张图片压缩成功")
                self._cp_show_success(f"压缩完成！共 {success_count} 张图片")

            self.cp_image_paths.clear()
            self._cp_refresh_grid()
        except Exception as e:
            err_msg = f"{e}\n{traceback.format_exc()}"
            print(f"[CRASH] {err_msg}")
            try:
                self._cp_log_message(f"致命错误: {err_msg}", is_error=True)
            except Exception:
                pass
            self._cp_hide_loading()
            QMessageBox.critical(self, "错误", err_msg)

    def _cp_log_message(self, message, is_error=False):
        prefix = "[错误] " if is_error else "[信息] "
        timestamp = datetime.now().strftime("%H:%M:%S")
        color = "#ff453a" if is_error else "#8e8e93"
        html = f'<span style="color:{color}">[{timestamp}] {prefix}{message}</span>'
        plain = f"[{timestamp}] {prefix}{message}"
        print(plain)
        try:
            self.cp_log_area.append(html)
            self.cp_log_area.ensureCursorVisible()
            QApplication.processEvents()
        except Exception as e:
            print(f"[压缩日志写入失败] {e}")

    def _on_cp_copy_log(self):
        text = self.cp_log_area.toPlainText()
        if text:
            QApplication.clipboard().setText(text)
            dlg = QMessageBox(self)
            dlg.setWindowTitle("提示")
            dlg.setText("日志已复制到剪贴板")
            dlg.setIcon(QMessageBox.NoIcon)
            dlg.setStandardButtons(QMessageBox.Ok)
            dlg.setFixedWidth(120)
            dlg.exec()

    # ── Watermark Interaction ──

    def _load_rotated_pixmap(self, path, target_width=180):
        """Load image with EXIF rotation applied, scale to target_width keeping aspect ratio."""
        try:
            pil_img = Image.open(path)
            pil_img = ImageOps.exif_transpose(pil_img)
            pil_img = pil_img.convert("RGBA")
            w = target_width
            ratio = pil_img.height / pil_img.width
            h = int(w * ratio)
            pil_img = pil_img.resize((w, h), Image.LANCZOS)
            data = pil_img.tobytes("raw", "RGBA")
            qimg = QImage(data, w, h, QImage.Format_RGBA8888)
            return QPixmap.fromImage(qimg.copy())
        except Exception:
            pixmap = QPixmap(path)
            if not pixmap.isNull():
                return pixmap.scaled(target_width, target_width * 2, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            return QPixmap()

    def _on_wm_select_images(self):
        dlg = ImagePickerDialog(self, self.wm_image_paths)
        if dlg.exec() == QDialog.Accepted:
            for p in dlg.get_selected_paths():
                if p not in self.wm_image_paths:
                    self.wm_image_paths.append(p)
            self._wm_refresh_grid()

    def _on_wm_clear_images(self):
        self.wm_image_paths.clear()
        self._wm_refresh_grid()

    def _wm_refresh_grid(self):
        # Clear old items
        for path, item in self.wm_image_items.items():
            item.deleteLater()
        self.wm_image_items.clear()
        for item in self._wm_skeleton_cards.values():
            item.deleteLater()
        self._wm_skeleton_cards.clear()
        self._wm_load_index = 0

        # Toggle empty state vs grid
        has_images = len(self.wm_image_paths) > 0
        self.wm_empty_state.setVisible(not has_images)
        self.wm_image_scroll.setVisible(has_images)
        self.wm_clear_btn.setVisible(has_images)

        if not has_images:
            return

        # Calculate thumbnail width
        scroll_width = self.wm_image_scroll.width()
        if scroll_width < 100:
            scroll_width = self.width() - 50
        spacing_total = 10 * 3
        padding_total = 8 * 4
        thumb_width = max(80, (scroll_width - spacing_total - padding_total) // 4)
        self._wm_thumb_width = thumb_width

        # Create skeleton cards for all images immediately
        for i, path in enumerate(self.wm_image_paths):
            row, col = i // 4, i % 4
            skeleton = self._wm_create_skeleton_card(thumb_width)
            self._wm_skeleton_cards[i] = skeleton
            self.wm_image_grid_layout.addWidget(skeleton, row, col)

        # Set initial scroll height
        ROW_HEIGHT = 250
        GRID_SPACING = 10
        PEEK_HEIGHT = ROW_HEIGHT // 4
        total_rows = (len(self.wm_image_paths) + 3) // 4
        if total_rows <= 2:
            scroll_h = total_rows * ROW_HEIGHT + (total_rows - 1) * GRID_SPACING
        else:
            scroll_h = 2 * ROW_HEIGHT + GRID_SPACING + PEEK_HEIGHT
        self.wm_image_scroll.setMinimumHeight(scroll_h)
        self.wm_image_scroll.setMaximumHeight(scroll_h)

        # Start replacing skeletons with real images one by one
        QTimer.singleShot(50, self._wm_load_next_image)

    def _wm_create_skeleton_card(self, tw):
        card = QFrame()
        card.setObjectName("wm_skeleton")
        card.setMinimumHeight(250)
        card.setMaximumWidth(tw + 8)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(4, 4, 4, 4)
        card_layout.setSpacing(2)
        # Fake delete row
        top_row = QWidget()
        top_layout = QHBoxLayout(top_row)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.addStretch()
        card_layout.addWidget(top_row)
        # Shimmer placeholder
        shimmer = QLabel()
        shimmer.setObjectName("wm_skeleton_shimmer")
        shimmer.setFixedSize(tw, 120)
        card_layout.addWidget(shimmer, 1)
        # Fake filename bar
        card_layout.addWidget(QLabel())
        return card

    def _wm_load_next_image(self):
        if self._wm_load_index >= len(self.wm_image_paths):
            return

        path = self.wm_image_paths[self._wm_load_index]
        index = self._wm_load_index
        row = index // 4
        col = index % 4
        tw = self._wm_thumb_width

        # Remove skeleton card at this position
        skeleton = self._wm_skeleton_cards.get(index)
        if skeleton:
            self.wm_image_grid_layout.removeWidget(skeleton)
            skeleton.deleteLater()
            del self._wm_skeleton_cards[index]

        # Create real card
        item = QFrame()
        item.setObjectName("wm_image_item")
        item.setMinimumHeight(250)
        item.setMaximumWidth(tw + 8)
        item_layout = QVBoxLayout(item)
        item_layout.setContentsMargins(4, 4, 4, 4)
        item_layout.setSpacing(2)

        # Top row: delete button at right
        top_row = QWidget()
        top_layout = QHBoxLayout(top_row)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.addStretch()

        del_btn = QPushButton("✕")
        del_btn.setObjectName("wm_delete_btn")
        del_btn.setCursor(Qt.PointingHandCursor)
        del_btn.clicked.connect(lambda checked=False, p=path: self._on_wm_delete_image(p))
        top_layout.addWidget(del_btn)
        item_layout.addWidget(top_row)

        # Thumbnail (clickable for preview)
        pixmap = self._load_rotated_pixmap(path, target_width=tw)
        thumb_btn = QPushButton()
        thumb_btn.setObjectName("wm_img_btn")
        thumb_btn.setCursor(Qt.PointingHandCursor)
        if not pixmap.isNull():
            scaled = pixmap.scaled(tw, 300, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            thumb_btn.setIcon(QIcon(scaled))
            thumb_btn.setIconSize(scaled.size())
            thumb_btn.setFixedSize(scaled.size())
        else:
            thumb_btn.setText("无法加载")
            thumb_btn.setFixedSize(tw, 200)
        thumb_btn.clicked.connect(lambda checked=False, p=path: self._on_wm_preview_image(p))
        item_layout.addWidget(thumb_btn, 1)

        # Filename label
        name_label = QLabel(os.path.basename(path))
        name_label.setObjectName("wm_filename_label")
        name_label.setAlignment(Qt.AlignCenter)
        item_layout.addWidget(name_label)

        self.wm_image_grid_layout.addWidget(item, row, col)
        self.wm_image_items[path] = item

        self._wm_load_index += 1
        QTimer.singleShot(0, self._wm_load_next_image)

    def _on_wm_delete_image(self, path):
        if path in self.wm_image_paths:
            self.wm_image_paths.remove(path)
            self._wm_refresh_grid()

    def _on_wm_preview_image(self, path):
        pixmap = self._load_rotated_pixmap(path, target_width=800)
        if pixmap.isNull():
            QMessageBox.warning(self, "提示", "无法加载图片")
            return
        # Scale to fit within 800x600 bounds
        h_ratio = 600 / pixmap.height() if pixmap.height() > 600 else 1
        w_ratio = 800 / pixmap.width() if pixmap.width() > 800 else 1
        scale = min(h_ratio, w_ratio)
        if scale < 1:
            pixmap = pixmap.scaled(int(pixmap.width() * scale), int(pixmap.height() * scale), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        dlg = QDialog(self)
        dlg.setWindowTitle(os.path.basename(path))
        dlg.setObjectName("wm_preview_dialog")
        dlg.setStyleSheet("QDialog#wm_preview_dialog { background: #1d1d1f; }")
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(8, 8, 8, 8)
        label = QLabel()
        label.setPixmap(pixmap)
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)
        dlg.exec()

    def _on_wm_all_cities_toggled(self, state):
        checked = bool(state)
        for cb in self.wm_city_checkboxes:
            cb.blockSignals(True)
            cb.setChecked(checked)
            cb.blockSignals(False)

    def _on_wm_city_checkbox_changed(self):
        all_checked = all(cb.isChecked() for cb in self.wm_city_checkboxes)
        self.wm_all_city_checkbox.blockSignals(True)
        self.wm_all_city_checkbox.setChecked(all_checked)
        self.wm_all_city_checkbox.blockSignals(False)

    def _on_wm_start(self):
        try:
            self.loading_overlay.show_loading()
            QApplication.processEvents()
            self._on_wm_start_inner()
        except Exception as e:
            err_msg = f"{e}\n{traceback.format_exc()}"
            print(f"[CRASH] {err_msg}")
            try:
                self.wm_log_area.clear()
                self.wm_log_area.append(f'<span style="color:#ff453a">致命错误: {err_msg}</span>')
            except Exception:
                pass
            self.loading_overlay.hide()
            QMessageBox.critical(self, "错误", err_msg)

    def _on_wm_start_inner(self):
        # Test log area first
        self.wm_log_area.clear()
        self._wm_log_message("水印任务启动...")
        QApplication.processEvents()

        images = self.wm_image_paths
        if not images:
            QMessageBox.warning(self, "提示", "请先选择图片")
            return

        selected_cities = [cb.text() for cb in self.wm_city_checkboxes if cb.isChecked()]
        if not selected_cities:
            QMessageBox.warning(self, "提示", "请至少选择一个城市")
            return

        year = self.wm_year_input.text().strip()
        month = self.wm_month_input.text().strip()
        if not year or not month:
            QMessageBox.warning(self, "提示", "请填写年份和月份")
            return
        try:
            year_int = int(year)
            month_int = int(month)
        except ValueError:
            QMessageBox.warning(self, "提示", "年份和月份必须是数字")
            return

        # Build address pool: round-robin across selected cities
        try:
            city_data = json.load(open(self._get_city_json_path(), encoding='utf-8'))
        except Exception as e:
            QMessageBox.warning(self, "提示", f"无法读取城市数据: {e}")
            return
        city_address_queues = {}
        for city in selected_cities:
            city_address_queues[city] = [item['sampleAddres'] for item in city_data.get(city, [])]
        address_pool = []
        if city_address_queues:
            max_len = max(len(q) for q in city_address_queues.values())
            for i in range(max_len):
                for city in selected_cities:
                    q = city_address_queues[city]
                    if i < len(q):
                        address_pool.append(q[i])

        # Date generation rules
        now = datetime.now()
        if month_int < now.month:
            max_day = calendar.monthrange(year_int, month_int)[1]
        elif month_int == now.month:
            max_day = now.day
        else:
            max_day = calendar.monthrange(year_int, month_int)[1]

        # Output folder
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        output_dir = os.path.join(desktop, "水印相册")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Font paths — try multiple options for packaged and dev environments
        font_candidates = [
            '/System/Library/Fonts/PingFang.ttc',
            '/System/Library/Fonts/STHeiti Light.ttc',
            '/Library/Fonts/Arial Unicode.ttf',
        ]
        font_path = None
        for fp in font_candidates:
            if os.path.exists(fp):
                font_path = fp
                break
        if not font_path:
            QMessageBox.warning(self, "提示", "无法找到可用的中文字体")
            return
        font_large = ImageFont.truetype(font_path, 29)
        font_small = ImageFont.truetype(font_path, 14)
        font_camera = ImageFont.truetype(font_path, 14)

        # Resolve icon_map.png path (for location pin icon)
        icon_map_candidates = [
            os.path.join(sys._MEIPASS, 'icon_map.png') if getattr(sys, 'frozen', False) else os.path.join(os.path.dirname(os.path.abspath(__file__)), 'icon_map.png'),
        ]
        icon_map_path = None
        for ip in icon_map_candidates:
            if os.path.exists(ip):
                icon_map_path = ip
                break

        # Weekday names
        weekday_names = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']

        self._wm_log_message(f"开始为 {len(images)} 张图片添加水印...")
        self._wm_log_message(f"选中城市: {', '.join(selected_cities)}")
        self._wm_log_message(f"日期范围: {year_int}年{month_int}月1日 - {year_int}年{month_int}月{max_day}日")
        self._wm_log_message(f"输出目录: {output_dir}")
        self._wm_log_message(f"字体: {font_path}")

        self._wm_log_message(f"地址池共 {len(address_pool)} 条, 开始处理 {len(images)} 张图片")

        success_count = 0
        for idx, img_path in enumerate(images):
            try:
                self._wm_log_message(f"[{idx+1}/{len(images)}] 正在处理: {os.path.basename(img_path)}")
                QApplication.processEvents()

                pil_img = Image.open(img_path)
                pil_img = ImageOps.exif_transpose(pil_img)
                pil_img = pil_img.convert("RGBA")

                w, h = pil_img.size
                self._wm_log_message(f"  图片尺寸: {w}×{h}")

                # ── Scale factor ──
                scale = h / 1600.0
                self._wm_log_message(f"  scale={scale:.2f}")
                font_time_pt = max(64, int(156 * scale))   # 时间字号1.35倍(原116pt×1.35)
                font_group_pt = max(24, int(42 * scale))    # 日期组不变
                font_cam_pt = max(26, int(45 * scale))      # 水印相机0.8倍(原56pt×0.8)
                font_time = ImageFont.truetype(font_path, font_time_pt)
                font_group = ImageFont.truetype(font_path, font_group_pt, index=4)  # PingFang SC Semibold 600w
                font_date = ImageFont.truetype(font_path, font_group_pt, index=3)  # PingFang SC Medium 500w (日期)
                font_cam = ImageFont.truetype(font_path, font_cam_pt, index=4)  # PingFang SC Semibold 600w
                self._wm_log_message(f"  字号: 时间={font_time_pt}pt, 日期组={font_group_pt}pt, 水印相机={font_cam_pt}pt")

                overlay = Image.new("RGBA", (w, h), (0, 0, 0, 0))
                draw = ImageDraw.Draw(overlay)

                # Generate random date and time
                random_day = random.randint(1, max_day)
                random_date = datetime(year_int, month_int, random_day)
                random_hour = random.randint(9, 18)
                random_minute = random.randint(0, 59)
                time_str = f"{random_hour:02d}:{random_minute:02d}"
                date_str = f"{year_int}.{month_int:02d}.{random_day:02d}"
                weekday_str = weekday_names[random_date.weekday()]

                # Get address for this image
                addr = address_pool[idx % len(address_pool)] if address_pool else ""

                # ── 1. Time (x居中, y=88.75%, 29pt, #FEFDFD) ──
                time_color = (0xFE, 0xFD, 0xFD, 255)
                time_y = int(h * 0.775) - int(40 * scale)
                time_bbox = draw.textbbox((0, 0), time_str, font=font_time)
                time_w = time_bbox[2] - time_bbox[0]
                time_x = (w - time_w) // 2
                draw.text((time_x, time_y), time_str, fill=time_color, font=font_time)
                self._wm_log_message(f"  时间: {time_str}, y={time_y}, x={time_x}, 字宽={time_w}, x居中")

                # ── 2+3. Date+Weekday+pin+address (时间底部 + 40pt间距) ──
                time_bbox = draw.textbbox((0, 0), time_str, font=font_time)
                # 时间文字底部 = time_y + bbox底部偏移
                time_bottom = time_y + time_bbox[3]
                # 25pt间距(日期组往上移15pt，原40pt)
                group_y = time_bottom + int(25 * scale)
                date_color = (0xFE, 0xFD, 0xFD, 255)   # 跟时间颜色一样
                wd_color = (0xFE, 0xFD, 0xFD, 255)     # 跟时间颜色一样
                addr_color = (0xFE, 0xFD, 0xFD, 255)   # 跟时间颜色一样

                # Measure widths
                date_bbox = draw.textbbox((0, 0), date_str, font=font_date)
                date_w = date_bbox[2] - date_bbox[0]
                wd_bbox = draw.textbbox((0, 0), weekday_str, font=font_group)
                wd_w = wd_bbox[2] - wd_bbox[0]
                addr_bbox = draw.textbbox((0, 0), addr, font=font_group) if addr else (0, 0, 0, 0)
                addr_w = addr_bbox[2] - addr_bbox[0] if addr else 0

                # Gaps: pin两侧间距与星期间距一致
                gap_date_wd = int(10 * scale)     # 日期距星期间距
                gap_wd_pin = int(10 * scale)      # pin距星期间距(=gap_date_wd)
                gap_pin_addr = int(10 * scale)    # 地址距pin间距(=gap_date_wd)
                # Pin icon: load PNG icon, size=48pt×scale
                pin_icon_size = max(32, int(48 * scale))
                pin_icon_img = Image.open(icon_map_path).convert('RGBA')
                pin_icon_img = pin_icon_img.resize((pin_icon_size, pin_icon_size), Image.LANCZOS)
                pin_icon_w = pin_icon_size
                # 定位图标顶部位置 = 时间底部 + 45pt×scale
                time_bottom = time_y + time_bbox[3]
                map_y = time_bottom + int(35 * scale)
                pin_y = map_y

                group_width = date_w + gap_date_wd + wd_w + gap_wd_pin + pin_icon_w + gap_pin_addr + addr_w
                group_x = (w - group_width) // 2

                # Draw date (ymd)
                draw.text((group_x, group_y), date_str, fill=date_color, font=font_date)

                # Draw weekday (5pt right of date)
                wd_x = group_x + date_w + gap_date_wd
                draw.text((wd_x, group_y), weekday_str, fill=wd_color, font=font_group)

                # Draw pin icon (PNG图标, 底部与日期文字底部对齐)
                pin_x = wd_x + wd_w + gap_wd_pin
                overlay.paste(pin_icon_img, (pin_x, pin_y), pin_icon_img)

                # Draw address (right of pin)
                if addr:
                    addr_x = pin_x + pin_icon_w + gap_pin_addr
                    draw.text((addr_x, group_y), addr, fill=addr_color, font=font_group)

                self._wm_log_message(f"  日期+星期+定位+地址: y={group_y}, 组x={group_x}, 组宽={group_width}")
                self._wm_log_message(f"    日期: {date_str}, x={group_x}, 宽={date_w}")
                self._wm_log_message(f"    星期: {weekday_str}, x={wd_x}, 宽={wd_w}")
                self._wm_log_message(f"    定位图标: pin_x={pin_x}, size={pin_icon_size}, pin_y={pin_y}")
                if addr:
                    self._wm_log_message(f"    地址: {addr}, x={addr_x}, 宽={addr_w}")

                # ── 4. 水印相机 (y=94%, 右边距=图片宽度4%, 微模糊) ──
                cam_color = (0xAE, 0xA3, 0x95, 255)
                cam_y = int(h * 0.94)
                cam_text = "水印相机"
                cam_bbox = draw.textbbox((0, 0), cam_text, font=font_cam)
                cam_w = cam_bbox[2] - cam_bbox[0]
                cam_x = w - int(w * 0.04) - cam_w
                # 水印相机单独图层 + 微模糊
                cam_layer = Image.new('RGBA', (w, h), (0, 0, 0, 0))
                cam_draw = ImageDraw.Draw(cam_layer)
                cam_draw.text((cam_x, cam_y), cam_text, fill=cam_color, font=font_cam)
                cam_layer = cam_layer.filter(ImageFilter.GaussianBlur(radius=max(1, int(0.8 * scale))))
                overlay = Image.alpha_composite(overlay, cam_layer)
                self._wm_log_message(f"  水印相机: y={cam_y}, x={cam_x}, 字宽={cam_w}, 右边距={int(w*0.04)}px")

                # Merge overlay onto original
                result = Image.alpha_composite(pil_img, overlay)
                result = result.convert("RGB")

                # Save
                basename = os.path.basename(img_path)
                name, ext = os.path.splitext(basename)
                out_path = os.path.join(output_dir, f"{name}_水印{ext}")
                result.save(out_path, quality=95)
                success_count += 1
                self._wm_log_message(f"  已保存: {out_path}")
                QApplication.processEvents()

            except Exception as e:
                self._wm_log_message(f"处理失败 {os.path.basename(img_path)}: {e}", is_error=True)
                self._wm_log_message(traceback.format_exc(), is_error=True)
                QApplication.processEvents()

        # Done
        msg = f"水印相册已生成，共 {success_count}/{len(images)} 张\n路径: {output_dir}"
        self._wm_log_message(msg)
        self.loading_overlay.show_success("水印相册已生成！")

    def _get_city_json_path(self):
        if getattr(sys, 'frozen', False):
            base = sys._MEIPASS
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, "city_address.json")

    def _show_msg(self, text):
        dlg = QMessageBox(self)
        dlg.setWindowTitle("提示")
        dlg.setText(text)
        dlg.setIcon(QMessageBox.NoIcon)
        dlg.setStandardButtons(QMessageBox.Ok)
        dlg.exec()

    # ── Watermark Log ──

    def _wm_log_message(self, message, is_error=False):
        prefix = "[错误] " if is_error else "[信息] "
        timestamp = datetime.now().strftime("%H:%M:%S")
        color = "#ff453a" if is_error else "#8e8e93"
        html = f'<span style="color:{color}">[{timestamp}] {prefix}{message}</span>'
        # Also print to console for debugging
        plain = f"[{timestamp}] {prefix}{message}"
        print(plain)
        try:
            self.wm_log_area.append(html)
            self.wm_log_area.ensureCursorVisible()
            QApplication.processEvents()
        except Exception as e:
            print(f"[水印日志写入失败] {e}")

    def _wm_flush_logs(self):
        # No longer needed for watermark — logs are written directly
        if self._wm_post_log_callback:
            callback = self._wm_post_log_callback
            self._wm_post_log_callback = None
            callback()

    def _on_wm_copy_log(self):
        text = self.wm_log_area.toPlainText()
        if text:
            QApplication.clipboard().setText(text)
            dlg = QMessageBox(self)
            dlg.setWindowTitle("提示")
            dlg.setText("日志已复制到剪贴板")
            dlg.setIcon(QMessageBox.NoIcon)
            dlg.setStandardButtons(QMessageBox.Ok)
            dlg.setFixedWidth(120)
            dlg.exec()

    # ── Excel Interaction ──

    def _on_select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx)"
        )
        if file_path:
            self.selected_file_path = file_path
            self.file_path_display.setText(file_path)
            self.excel_name_input.setEnabled(False)
            self.excel_name_input.clear()
            self.file_clear_btn.setVisible(True)

    def _on_clear_file(self):
        self.selected_file_path = ""
        self.file_path_display.clear()
        self.excel_name_input.setEnabled(True)
        self.excel_name_input.setText("订单表")
        self.file_clear_btn.setVisible(False)

    def _on_excel_name_changed(self, text):
        pass

    def _on_copy_log(self):
        text = self.log_area.toPlainText()
        if text:
            from PySide6.QtGui import QClipboard
            clipboard = QApplication.clipboard()
            clipboard.setText(text)
            dlg = QMessageBox(self)
            dlg.setWindowTitle("提示")
            dlg.setText("复制日志成功")
            dlg.setIcon(QMessageBox.NoIcon)
            dlg.setStandardButtons(QMessageBox.Ok)
            dlg.setFixedWidth(120)
            dlg.exec()

    def log_message(self, message, is_error=False):
        prefix = "[错误] " if is_error else "[信息] "
        timestamp = datetime.now().strftime("%H:%M:%S")
        color = "#ff453a" if is_error else "#8e8e93"
        self._pending_logs.append(
            f'<span style="color:{color}">[{timestamp}] {prefix}{message}</span>'
        )

    def _flush_logs(self):
        if self._pending_logs:
            msg = self._pending_logs.pop(0)
            self.log_area.append(msg)
            QTimer.singleShot(100, self._flush_logs)
        elif self._post_log_callback:
            callback = self._post_log_callback
            self._post_log_callback = None
            callback()

    def _show_success_dialog(self, text):
        dlg = QMessageBox(self)
        dlg.setWindowTitle("成功")
        dlg.setText(text)
        dlg.setIcon(QMessageBox.NoIcon)
        dlg.setStandardButtons(QMessageBox.Ok)
        dlg.exec()

    # ─── Business Logic (UNCHANGED) ─────────────────────────────────────────────

    def generate_random_datetime(self, month):
        now = datetime.now()
        year = now.year
        max_day = calendar.monthrange(year, month)[1]
        if month == now.month:
            max_day = now.day
        day = random.randint(1, max_day)
        base = datetime(year, month, day, 9, 0, 0)
        random_seconds = random.randint(0, 43200)
        result = base + timedelta(seconds=random_seconds)
        return result.strftime("%Y-%m-%d %H:%M:%S")

    def generate_order_number(self, month):
        now = datetime.now()
        year = now.year
        max_day = calendar.monthrange(year, month)[1]
        if month == now.month:
            max_day = now.day
        day = random.randint(1, max_day)
        random_date = datetime(year, month, day)
        yymmdd = random_date.strftime("%y%m%d")
        random_part = "1" + "".join([str(random.randint(0, 9)) for _ in range(13)])
        return f"S{yymmdd}{random_part}"

    def generate_order_number_from_date(self, date_str):
        dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        yymmdd = dt.strftime("%y%m%d")
        random_part = "1" + "".join([str(random.randint(0, 9)) for _ in range(13)])
        return f"S{yymmdd}{random_part}"

    def generate_pos_number_from_date(self, date_str):
        dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        yyyymmdd = dt.strftime("%Y%m%d")
        random_part = "".join([str(random.randint(0, 9)) for _ in range(5)])
        return f"OS{yyyymmdd}-{random_part}"

    def find_column_by_header(self, ws, header_name):
        for col in range(1, ws.max_column + 2):
            if ws.cell(row=1, column=col).value == header_name:
                return col
        return None

    def find_first_empty_column(self, ws):
        for col in range(1, ws.max_column + 2):
            val = ws.cell(row=1, column=col).value
            if val is None or val == "":
                return col
        return ws.max_column + 1

    def get_last_data_row(self, ws, col):
        for row in range(ws.max_row, 0, -1):
            val = ws.cell(row=row, column=col).value
            if val is not None and val != "":
                return row
        return 0

    def count_data_cells(self, ws, col):
        count = 0
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and val != "":
                count += 1
        return count

    def _sort_date_column(self, ws, date_col):
        last_row = ws.max_row
        if last_row <= 1:
            return
        dates = []
        for row in range(2, last_row + 1):
            dates.append(ws.cell(row=row, column=date_col).value)
        dates.sort(
            key=lambda d: (1, d or '') if d else (0, ''),
            reverse=True
        )
        for i, date_val in enumerate(dates):
            row = i + 2
            if date_val is not None:
                cell = ws.cell(row=row, column=date_col, value=date_val)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws.cell(row=row, column=date_col).value = None

    def start_generation(self):
        self._pending_logs.clear()
        self._post_log_callback = None
        # === 校验Excel名称 ===
        excel_name = self.excel_name_input.text().strip()
        if not excel_name and not self.selected_file_path:
            self.log_message("Excel名称为空且未选择文件", is_error=True)
            self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "请填写Excel名称或选择已有Excel文件")
            self._flush_logs()
            return

        # === 校验月份 ===
        month_str = self.month_input.text().strip()
        if not month_str:
            self.log_message("月份为空", is_error=True)
            self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "请填写月份(1-12)")
            self._flush_logs()
            return
        try:
            month = int(month_str)
            if month < 1 or month > 12:
                self.log_message("月份不在1-12范围", is_error=True)
                self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "月份必须在1-12之间")
                self._flush_logs()
                return
        except ValueError:
            self.log_message("月份格式错误", is_error=True)
            self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "月份必须为整数")
            self._flush_logs()
            return

        # === 校验销售日期总条数 ===
        date_count_str = self.date_count_input.text().strip()
        if date_count_str:
            try:
                date_count = int(date_count_str)
                if date_count <= 0:
                    self.log_message("销售日期总条数必须大于0", is_error=True)
                    self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "销售日期总条数必须大于0")
                    self._flush_logs()
                    return
            except ValueError:
                self.log_message("销售日期总条数格式错误", is_error=True)
                self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "销售日期总条数必须为整数")
                self._flush_logs()
                return
        else:
            date_count = 0

        # === 校验销售单号总条数 ===
        order_count_str = self.order_count_input.text().strip()
        if order_count_str:
            try:
                order_count = int(order_count_str)
                if order_count <= 0:
                    self.log_message("销售单号总条数必须大于0", is_error=True)
                    self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "销售单号总条数必须大于0")
                    self._flush_logs()
                    return
            except ValueError:
                self.log_message("销售单号总条数格式错误", is_error=True)
                self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "销售单号总条数必须为整数")
                self._flush_logs()
                return
        else:
            order_count = 0

        # === 校验Pos单号总条数 ===
        pos_count_str = self.pos_count_input.text().strip()
        if pos_count_str:
            try:
                pos_count = int(pos_count_str)
                if pos_count <= 0:
                    self.log_message("Pos单号总条数必须大于0", is_error=True)
                    self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "Pos单号总条数必须大于0")
                    self._flush_logs()
                    return
            except ValueError:
                self.log_message("Pos单号总条数格式错误", is_error=True)
                self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "Pos单号总条数必须为整数")
                self._flush_logs()
                return
        else:
            pos_count = 0

        # 至少一个总条数大于0
        if date_count <= 0 and order_count <= 0 and pos_count <= 0:
            self.log_message("三个总条数均为0", is_error=True)
            self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "销售日期、销售单号和Pos单号总条数至少需要一个大于0")
            self._flush_logs()
            return

        # 确定文件路径
        if self.selected_file_path:
            file_path = self.selected_file_path
        else:
            if not excel_name.endswith(".xlsx"):
                excel_name += ".xlsx"
            base_dir = os.path.expanduser("~/Desktop")
            file_path = os.path.join(base_dir, excel_name)

        try:
            # === 加载或创建Excel ===
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
                self.log_message(f"加载已有文件: {file_path}")
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "订单数据"
                self.log_message(f"创建新文件: {file_path}")

            # === 处理销售日期列 ===
            date_col = None
            if date_count > 0:
                date_col = self.find_column_by_header(ws, "销售日期")
                if date_col is not None:
                    last_row = self.get_last_data_row(ws, date_col)
                    date_start_row = last_row + 1
                    self.log_message(f"销售日期列已存在(列{date_col})，追加{date_count}条数据")
                else:
                    date_col = self.find_first_empty_column(ws)
                    ws.cell(row=1, column=date_col, value="销售日期")
                    date_start_row = 2
                    self.log_message(f"在列{date_col}新建销售日期标题")

                dates = [self.generate_random_datetime(month) for _ in range(date_count)]
                for i, dt in enumerate(dates):
                    cell = ws.cell(row=date_start_row + i, column=date_col, value=dt)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                self.log_message(f"写入{date_count}条销售日期")

            # 如果只生成单号（没有新日期），必须找到已有日期列
            if date_count <= 0 and order_count > 0:
                date_col = self.find_column_by_header(ws, "销售日期")
                if date_col is None:
                    self.log_message("没有销售日期列，无法生成单号", is_error=True)
                    self._post_log_callback = lambda: QMessageBox.warning(self, "提示", "无法生成销售单号：文件中没有销售日期数据")
                    self._flush_logs()
                    return
                self.log_message(f"使用已有销售日期列(列{date_col})")

            # === 排序前先统计已有单号条数 ===
            existing_order_col = self.find_column_by_header(ws, "销售单号")
            existing_order_length = 0
            if existing_order_col is not None:
                existing_order_length = self.count_data_cells(ws, existing_order_col)

            # === 仅排序销售日期列 ===
            if date_col is not None:
                self._sort_date_column(ws, date_col)
                self.log_message("已按销售日期降序排序（仅日期列）")

            # === 计算总日期长度 ===
            total_date_length = 0
            if date_col is not None:
                total_date_length = self.count_data_cells(ws, date_col)
                self.log_message(f"销售日期总长度: {total_date_length}条")

            # === 处理销售单号列 ===
            order_col = None
            actual_order_count = 0

            if order_count > 0:
                if existing_order_col is not None:
                    order_col = existing_order_col
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=order_col).value = None
                    a = existing_order_length + order_count
                    if a >= total_date_length:
                        actual_order_count = total_date_length
                    else:
                        actual_order_count = a
                    self.log_message(
                        f"已有单号{existing_order_length}条+新增{order_count}={a}，"
                        f"实际重新生成{actual_order_count}条单号"
                    )
                else:
                    order_col = date_col + 1 if date_col else self.find_first_empty_column(ws)
                    next_val = ws.cell(row=1, column=order_col).value
                    if next_val is not None and next_val != "":
                        ws.insert_cols(order_col)
                        self.log_message(f"在列{order_col}插入新列(原有数据右移)")
                    ws.cell(row=1, column=order_col, value="销售单号")
                    actual_order_count = min(order_count, total_date_length)
                    self.log_message(
                        f"新建单号列(列{order_col})，生成{actual_order_count}条"
                        f"(请求{order_count}条，日期长度{total_date_length}条)"
                    )

                for i in range(actual_order_count):
                    row = 2 + i
                    date_value = ws.cell(row=row, column=date_col).value if date_col else None
                    if date_value:
                        order_num = self.generate_order_number_from_date(date_value)
                    else:
                        order_num = self.generate_order_number(month)
                    cell = ws.cell(row=row, column=order_col, value=order_num)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                self.log_message(f"已从上到下重新生成{actual_order_count}条销售单号，单号日期与同行日期一致")

            # === 处理Pos单号列 ===
            pos_col = None
            actual_pos_count = 0

            if pos_count > 0:
                existing_pos_col = self.find_column_by_header(ws, "Pos单号")
                existing_pos_length = 0
                if existing_pos_col is not None:
                    existing_pos_length = self.count_data_cells(ws, existing_pos_col)

                if existing_pos_col is not None:
                    pos_col = existing_pos_col
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=pos_col).value = None
                    a = existing_pos_length + pos_count
                    if a >= total_date_length:
                        actual_pos_count = total_date_length
                    else:
                        actual_pos_count = a
                    self.log_message(
                        f"已有Pos单号{existing_pos_length}条+新增{pos_count}={a}，"
                        f"实际重新生成{actual_pos_count}条Pos单号"
                    )
                else:
                    pos_col = (order_col or date_col) + 1 if (order_col or date_col) else self.find_first_empty_column(ws)
                    next_val = ws.cell(row=1, column=pos_col).value
                    if next_val is not None and next_val != "":
                        ws.insert_cols(pos_col)
                        self.log_message(f"在列{pos_col}插入新列(原有数据右移)")
                    ws.cell(row=1, column=pos_col, value="Pos单号")
                    actual_pos_count = min(pos_count, total_date_length)
                    self.log_message(
                        f"新建Pos单号列(列{pos_col})，生成{actual_pos_count}条"
                        f"(请求{pos_count}条，日期长度{total_date_length}条)"
                    )

                for i in range(actual_pos_count):
                    row = 2 + i
                    date_value = ws.cell(row=row, column=date_col).value if date_col else None
                    if date_value:
                        pos_num = self.generate_pos_number_from_date(date_value)
                    else:
                        pos_num = self.generate_pos_number_from_date(
                            self.generate_random_datetime(month)
                        )
                    cell = ws.cell(row=row, column=pos_col, value=pos_num)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                self.log_message(f"已从上到下重新生成{actual_pos_count}条Pos单号，Pos单号日期与同行日期一致")

            # 调整列宽
            if date_col is not None:
                ws.column_dimensions[get_column_letter(date_col)].width = 22
            if order_col is not None:
                ws.column_dimensions[get_column_letter(order_col)].width = 25
            if pos_col is not None:
                ws.column_dimensions[get_column_letter(pos_col)].width = 22

            wb.save(file_path)
            self.log_message(f"文件保存成功: {file_path}")
            self._post_log_callback = lambda: self._show_success_dialog(f"Excel文件已生成!\n路径: {file_path}")
            self._flush_logs()

        except PermissionError:
            msg = "文件保存失败: 可能被其他程序占用，请关闭Excel后重试"
            self.log_message(f"{msg}\n路径: {file_path}", is_error=True)
            self._post_log_callback = lambda: QMessageBox.critical(self, "错误", msg)
            self._flush_logs()
        except Exception as e:
            msg = f"生成失败: {str(e)}"
            self.log_message(f"{msg}\n{traceback.format_exc()}", is_error=True)
            self._post_log_callback = lambda: QMessageBox.critical(self, "错误", msg)
            self._flush_logs()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelGeneratorApp()
    window.show()
    sys.exit(app.exec())